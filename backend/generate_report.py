import json, os, sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage

# ── Load raw data & filter ──
with open('../march-2026-raw.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

raw_cols = data['cols']
raw_rows = data['rows']

loc_idx = raw_cols.index('location_name')
filtered = [r for r in raw_rows if r[loc_idx] == 'BYD AP Automotive Mukdahan']
paid_idx = raw_cols.index('paid_date_bkk')
filtered.sort(key=lambda r: r[paid_idx] or '')
print(f'Filtered rows: {len(filtered)}')

# ── Manual inputs ──
# Accept electricity bill image path from CLI arg
ELECTRICITY_BILL_IMAGE = sys.argv[1] if len(sys.argv) > 1 else None

ELECTRICITY_COST = 21527.28
INTERNET_COST = 598.00
ETAX = 184.00
TRANSFER_FEE = 30.00
TRANSACTION_FEE_RATE = 0.0365
VAT_RATE = 0.07
LOCATION_SHARE_RATE = 0.40
LOCATION_NAME = 'BYD AP Automotive Mukdahan'


def parse_dt(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace('Z', '').replace('+00:00', ''))
    except:
        return s


def map_row(r):
    invoice_id = r[raw_cols.index('invoice_id')]
    reference_id = r[raw_cols.index('reference_id')]
    location_name = r[raw_cols.index('location_name')]
    evse_name = r[raw_cols.index('evse_name')]
    session_start = r[raw_cols.index('session_start_bkk')]
    session_end = r[raw_cols.index('session_end_bkk')]
    paid_date = r[raw_cols.index('paid_date_bkk')]
    transaction_id = r[raw_cols.index('payment_transaction_id')]
    invoice_status = r[raw_cols.index('invoice_status')]
    payment_amount = r[raw_cols.index('payment_amount')]
    invoice_amount = r[raw_cols.index('invoice_amount')]
    total_discount = r[raw_cols.index('total_discount')]
    kwh = r[raw_cols.index('kwh')]
    unit_price = None
    if kwh and float(kwh) > 0 and invoice_amount:
        unit_price = float(invoice_amount) / float(kwh)
    total_time = r[raw_cols.index('total_time')]
    total_overtime = r[raw_cols.index('total_overtime')]
    etax_number = r[raw_cols.index('etax_number')]
    privilege_name = r[raw_cols.index('discount_label')]
    vin = r[raw_cols.index('vin')]
    return [
        invoice_id, reference_id, location_name, evse_name,
        parse_dt(session_start), parse_dt(session_end), parse_dt(paid_date),
        transaction_id, invoice_status, payment_amount,
        invoice_amount, total_discount, unit_price, kwh,
        total_time, total_overtime, etax_number, None,
        privilege_name, vin, None,
    ]


# ── Calculations ──
revenue = sum(float(r_[raw_cols.index('payment_amount')] or 0) for r_ in filtered)
transaction_fee = revenue * TRANSACTION_FEE_RATE
vat_on_fee = transaction_fee * VAT_RATE
total_fee = transaction_fee + vat_on_fee + TRANSFER_FEE
internet_incl_vat = INTERNET_COST * (1 + VAT_RATE)
etax_incl_vat = ETAX * (1 + VAT_RATE)
remaining = revenue - total_fee - ELECTRICITY_COST - internet_incl_vat - etax_incl_vat
location_share = remaining * LOCATION_SHARE_RATE
vat_portion = location_share - (location_share / (1 + VAT_RATE))
before_vat = location_share / (1 + VAT_RATE)

# ── Styles ──
thin = Side(style='thin')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_bottom = Border(bottom=thin)

font_header = Font(name='Calibri', size=11, bold=True)
font_data = Font(name='Calibri', size=11)
font_bold = Font(name='Calibri', size=11, bold=True)
font_bold_brown = Font(name='Calibri', size=11, bold=True, color='8B4513')
font_big_bold = Font(name='Calibri', size=13, bold=True)

fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
fill_light_blue = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')

align_right = Alignment(horizontal='right')
align_center = Alignment(horizontal='center')
num_2dp = '#,##0.00'


def write_summary_row(ws, row, label, desc, value, label_font=font_bold,
                      val_fill=None, val_font=None, blue_row=False,
                      border=None):
    c_b = ws.cell(row=row, column=2, value=label)
    c_b.font = label_font
    c_b.alignment = Alignment(horizontal='right')
    c_c = ws.cell(row=row, column=3, value=desc)
    c_c.alignment = align_center
    c_d = ws.cell(row=row, column=4, value=round(value, 2) if value is not None else None)
    c_d.number_format = num_2dp
    c_d.alignment = align_right
    if val_fill:
        c_d.fill = val_fill
    if val_font:
        c_d.font = val_font
    if blue_row:
        for c in (c_b, c_c, c_d):
            c.fill = fill_light_blue
            c.border = border_all
            c.font = font_bold_brown
    if border:
        c_d.border = border


# ═══════════════════════════════════════════════
# WORKBOOK
# ═══════════════════════════════════════════════
wb = Workbook()

# ── Summary Sheet ──
ws = wb.active
ws.title = 'Summary'
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 38
ws.column_dimensions['D'].width = 18

# Revenue header row with borders
for col in (2, 3, 4):
    ws.cell(row=2, column=col).border = border_all
ws.cell(row=2, column=2, value='Revenue').font = font_bold
ws.cell(row=2, column=2).alignment = align_center
ws.cell(row=2, column=4, value=round(revenue, 2)).font = font_bold
ws.cell(row=2, column=4).number_format = num_2dp
ws.cell(row=2, column=4).alignment = align_right

# Fee section
write_summary_row(ws, 4, 'Transaction Fee', '(3.65% of Revenue)', transaction_fee)
write_summary_row(ws, 5, 'VAT', '(7% of Transaction Fee)', vat_on_fee)
write_summary_row(ws, 6, 'Transfer', None, TRANSFER_FEE)
write_summary_row(ws, 7, 'Total Fee', None, total_fee, border=border_bottom)

# Cost section
write_summary_row(ws, 9, 'Electricity Cost', None, ELECTRICITY_COST,
                  val_fill=fill_yellow, val_font=font_bold)
write_summary_row(ws, 10, 'Internet Cost', None, INTERNET_COST)
write_summary_row(ws, 11, None, 'Vat 7%', internet_incl_vat)
write_summary_row(ws, 12, 'Etax', None, ETAX)
write_summary_row(ws, 13, 'Etax (Include Vat)', 'Vat 7%', etax_incl_vat)
write_summary_row(ws, 14, '\u0e04\u0e07\u0e40\u0e2b\u0e25\u0e37\u0e2d', None, remaining,
                  val_font=font_bold, border=border_bottom)

# Location share section (blue rows)
write_summary_row(ws, 16, LOCATION_NAME, f'({int(LOCATION_SHARE_RATE*100)}% of Total Revenue VAT Incl.)',
                  location_share, blue_row=True)
write_summary_row(ws, 17, 'VAT', '(7% of Cash In)', vat_portion, blue_row=True)
# Before VAT row
ws.cell(row=18, column=3, value='(Before VAT)').alignment = align_center
ws.cell(row=18, column=4, value=round(before_vat, 2)).number_format = num_2dp
ws.cell(row=18, column=4).alignment = align_right
for c in (2, 3, 4):
    ws.cell(row=18, column=c).fill = fill_light_blue
    ws.cell(row=18, column=c).border = border_all
    ws.cell(row=18, column=c).font = font_bold_brown

# Net GP
ws.cell(row=20, column=2, value='Net GP').font = font_big_bold
ws.cell(row=20, column=2).alignment = align_center
ws.cell(row=20, column=3, value='(VAT Included)').font = font_bold
ws.cell(row=20, column=3).alignment = align_center
ws.cell(row=20, column=4, value=round(location_share, 2)).font = font_big_bold
ws.cell(row=20, column=4).number_format = num_2dp
ws.cell(row=20, column=4).alignment = align_right
ws.cell(row=20, column=4).border = Border(
    top=Side(style='double'), bottom=Side(style='double'))

# ── Electricity bill image (next to summary table) ──
if ELECTRICITY_BILL_IMAGE and os.path.exists(ELECTRICITY_BILL_IMAGE):
    img = XlImage(ELECTRICITY_BILL_IMAGE)
    # Scale to fit nicely - approx 500px wide
    max_w = 500
    ratio = max_w / img.width if img.width > max_w else 1
    img.width = int(img.width * ratio)
    img.height = int(img.height * ratio)
    # Place at column F (6), row 2 - right of the summary table
    ws.add_image(img, 'F2')
    print(f'Embedded electricity bill image: {ELECTRICITY_BILL_IMAGE}')
else:
    if ELECTRICITY_BILL_IMAGE:
        print(f'WARNING: Image not found: {ELECTRICITY_BILL_IMAGE}')
    else:
        print('No electricity bill image provided (pass path as CLI arg)')


# ── Data Sheet ──
ws_data = wb.create_sheet('03.2026')

headers = [
    'Invoice', 'Reference Id', 'Location', 'Evse', 'Start Date Time', 'End Date Time',
    'Payment Created At', 'Transaction Id', 'Invoice Status', 'Payment Amount',
    'Total Cost', 'Total Discount', 'Unit Price\u00a0', 'Kwh', 'Total Time (Hour)',
    'Total Overtime (Hour)', 'Etax Number', None, 'Privilege Name', 'RFID Number',
    'organization_name'
]

for col_idx, h in enumerate(headers, 1):
    cell = ws_data.cell(row=1, column=col_idx, value=h)
    cell.font = font_header

for row_num, raw_row in enumerate(filtered, 2):
    mapped = map_row(raw_row)
    for col_idx, val in enumerate(mapped, 1):
        cell = ws_data.cell(row=row_num, column=col_idx, value=val)
        cell.font = font_data
        if col_idx in (5, 6, 7) and isinstance(val, datetime):
            cell.number_format = 'm/d/yy h:mm'
        elif col_idx in (10, 11, 12, 13, 14, 15, 16, 18):
            cell.number_format = '0.00'

# Auto-width
from openpyxl.utils import get_column_letter
for col_idx in range(1, len(headers) + 1):
    max_len = len(str(headers[col_idx-1] or ''))
    for r in range(2, min(20, ws_data.max_row + 1)):
        val = ws_data.cell(row=r, column=col_idx).value
        max_len = max(max_len, min(len(str(val or '')), 40))
    ws_data.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

# ── Save ──
output = os.path.join("C:", os.sep, "auto_report_system", "BYD AP Automotive Mukdahan 03.2026 v4.xlsx")
wb.save(output)

fsize = os.path.getsize(output)
print(f'Saved: {output} ({fsize:,} bytes)')
print(f'\n--- Summary ---')
print(f'Revenue:          {revenue:>12,.2f}')
print(f'Transaction Fee:  {transaction_fee:>12,.2f}')
print(f'VAT on Fee:       {vat_on_fee:>12,.2f}')
print(f'Transfer:         {TRANSFER_FEE:>12,.2f}')
print(f'Total Fee:        {total_fee:>12,.2f}')
print(f'Electricity:      {ELECTRICITY_COST:>12,.2f}')
print(f'Internet(+VAT):   {internet_incl_vat:>12,.2f}')
print(f'Etax(+VAT):       {etax_incl_vat:>12,.2f}')
print(f'Remaining:        {remaining:>12,.2f}')
print(f'BYD share (40%):  {location_share:>12,.2f}')
print(f'VAT portion:      {vat_portion:>12,.2f}')
print(f'Before VAT:       {before_vat:>12,.2f}')
print(f'Net GP:           {location_share:>12,.2f}')
