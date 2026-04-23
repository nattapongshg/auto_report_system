"""Parse PEA/MEA electricity bill exports into row dicts.

MEA: real .xlsx, first sheet, row 1 = header. Columns we care about:
     CA, KWH-TOT, Sub Total (pre-VAT), Vat Amt., Total (VAT-incl),
     Inv. No., Due Date.
     Dates are datetime objects with Buddhist year (e.g. 2569 → 2026).
     Bill period = month of (Due Date − 1 month).

PEA: '.xls' extension but actually tab-separated TIS-620 text with some
     preamble lines before a tab-separated header starting with 'Item\\t'.
     Columns we care about: CUST NO. (CA), UNIT (kWh), AMOUNT (VAT-incl),
     VAT, INVOICE NO, BILLPERIOD (e.g. '256903' → BE 2569 / month 03).
     Storage convention: total = PEA AMOUNT, vat = PEA VAT,
     amount = total - vat (pre-VAT base). This keeps semantics identical
     to MEA where amount=Sub Total (pre-VAT), vat=Vat Amt, total=Total.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import BinaryIO

import openpyxl


@dataclass
class BillRow:
    provider: str              # 'mea' | 'pea'
    ca: str
    year_month: str            # '2026-03'
    kwh: Decimal | None
    amount: Decimal | None     # pre-VAT base (for MEA); for PEA we only know total
    vat: Decimal | None
    total: Decimal             # VAT-inclusive — the number the report uses
    invoice_no: str | None
    bill_date: date | None
    raw: dict


def _dec(v: object) -> Decimal | None:
    if v is None or v == "":
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _pad_ca(v: object) -> str | None:
    """Normalize a CA to canonical form = numeric with no leading zeros.
    PEA bill exports include an extra leading 0 in the CA column (12 chars)
    that isn't part of the real account number; SHARGE LOCATIONS stores PEA
    CAs as plain ints. Stripping zeros on both sides keeps them aligned."""
    if v is None or v == "":
        return None
    s = str(v).strip()
    if not s:
        return None
    if s.isdigit():
        s = s.lstrip("0") or "0"
    return s


def _be_to_ce(y: int) -> int:
    return y - 543 if y > 2400 else y


def _minus_one_month(d: date) -> date:
    y, m = d.year, d.month - 1
    if m == 0:
        y, m = y - 1, 12
    return date(y, m, 1)


def parse_mea(data: bytes, source_file: str | None = None) -> list[BillRow]:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it)]
    col = {h: i for i, h in enumerate(header)}

    required = ["CA", "Total", "Due Date"]
    missing = [c for c in required if c not in col]
    if missing:
        raise ValueError(f"MEA file missing columns: {missing}")

    rows: list[BillRow] = []
    for r in it:
        if r is None or all(v is None or v == "" for v in r):
            continue
        ca = _pad_ca(r[col["CA"]])
        total = _dec(r[col["Total"]])
        if not ca or total is None:
            continue

        due_raw = r[col["Due Date"]]
        if isinstance(due_raw, datetime):
            due_ce = due_raw.replace(year=_be_to_ce(due_raw.year)).date()
        elif isinstance(due_raw, date):
            due_ce = due_raw.replace(year=_be_to_ce(due_raw.year))
        else:
            continue
        period = _minus_one_month(due_ce)
        year_month = f"{period.year:04d}-{period.month:02d}"

        rows.append(BillRow(
            provider="mea",
            ca=ca,
            year_month=year_month,
            kwh=_dec(r[col.get("KWH-TOT", -1)]) if "KWH-TOT" in col else None,
            amount=_dec(r[col.get("Sub Total", -1)]) if "Sub Total" in col else None,
            vat=_dec(r[col.get("Vat Amt.", -1)]) if "Vat Amt." in col else None,
            total=total,
            invoice_no=str(r[col["Inv. No."]]).strip() if "Inv. No." in col and r[col["Inv. No."]] else None,
            bill_date=due_ce,
            raw={"source": source_file, "header": header, "row": [str(v) if v is not None else None for v in r]},
        ))
    return rows


def _read_pea_tsv(data: bytes) -> tuple[list[str], list[list[str]]]:
    """Skip preamble until we find a line starting with 'Item\\t', then parse
    remaining as TSV."""
    text = data.decode("tis-620", errors="replace")
    lines = text.splitlines()
    # Find header row
    header_i = None
    for i, line in enumerate(lines):
        if line.startswith("Item\t"):
            header_i = i
            break
    if header_i is None:
        raise ValueError("PEA file: could not find 'Item' header row")

    reader = csv.reader(io.StringIO("\n".join(lines[header_i:])), delimiter="\t")
    header = [h.strip() for h in next(reader)]
    data_rows = [row for row in reader if any(c.strip() for c in row)]
    return header, data_rows


def parse_pea(data: bytes, source_file: str | None = None) -> list[BillRow]:
    header, data_rows = _read_pea_tsv(data)
    col = {h: i for i, h in enumerate(header)}

    required = ["CUST NO.", "AMOUNT", "BILLPERIOD"]
    missing = [c for c in required if c not in col]
    if missing:
        raise ValueError(f"PEA file missing columns: {missing}")

    rows: list[BillRow] = []
    for r in data_rows:
        if len(r) < len(header):
            r = r + [""] * (len(header) - len(r))
        ca = _pad_ca(r[col["CUST NO."]])
        raw_amount = _dec(r[col["AMOUNT"]])
        vat = _dec(r[col.get("VAT", -1)]) if "VAT" in col else None
        period = r[col["BILLPERIOD"]].strip() if r[col["BILLPERIOD"]] else ""
        if not ca or raw_amount is None or len(period) != 6:
            continue

        # PEA AMOUNT is VAT-inclusive. Map to the same {amount, vat, total}
        # shape MEA uses: total = VAT-incl, vat = VAT portion, amount = base.
        total = raw_amount
        amount = total - (vat or 0)

        be_year = int(period[:4])
        month = int(period[4:])
        year_month = f"{_be_to_ce(be_year):04d}-{month:02d}"
        bill_date = date(_be_to_ce(be_year), month, 1)

        rows.append(BillRow(
            provider="pea",
            ca=ca,
            year_month=year_month,
            kwh=_dec(r[col.get("UNIT", -1)]) if "UNIT" in col else None,
            amount=amount,
            vat=vat,
            total=total,
            invoice_no=r[col["INVOICE NO"]].strip() if "INVOICE NO" in col and r[col["INVOICE NO"]] else None,
            bill_date=bill_date,
            raw={"source": source_file, "header": header, "row": r},
        ))
    return rows


def detect_and_parse(filename: str, data: bytes) -> list[BillRow]:
    """Best-effort provider detection by file contents (MEA is real xlsx, PEA
    is TSV-in-xls). Falls back to filename hints."""
    # openpyxl/xlsx files start with 'PK' (zip magic).
    if data[:2] == b"PK":
        return parse_mea(data, source_file=filename)
    # PEA export is text (TIS-620).
    return parse_pea(data, source_file=filename)
