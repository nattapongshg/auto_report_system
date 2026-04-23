"""Safe formula evaluator + Summary-sheet renderer driven by report_layout_templates.

A template `summary_layout` is a list of row objects. Each row is rendered
against a context dict containing live variables — the operator can edit the
template's JSON to change what appears in the Summary sheet without touching
Python code.

Context variables always available:
  revenue, electricity_cost, internet_cost, etax, evse_count,
  location_name, group_name, date_start, date_end

Template params feed the context (tx_rate, vat_rate, wht_rate, transfer_fee,
per_evse_internet, ...). The per-location share_rate always wins over params.

Derived (computed before eval):
  sharge_rate = 1 - share_rate                         (for dealer)
  location_share                                        (filled in share rows)
  total_payment_to_sharge, total_payment_to_dealer      (dealer template)

Formulas are restricted: only arithmetic + names already in context. No
attribute access, no calls except a small allowlist (min, max, round, abs).
"""
from __future__ import annotations

import ast
import operator as _op
import re
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.engine.share_calc import compute_totals


# ── Safe evaluator ───────────────────────────────────────────────
_BIN_OPS = {
    ast.Add: _op.add, ast.Sub: _op.sub, ast.Mult: _op.mul,
    ast.Div: _op.truediv, ast.Mod: _op.mod, ast.Pow: _op.pow,
    ast.FloorDiv: _op.floordiv,
}
_UNARY_OPS = {ast.UAdd: _op.pos, ast.USub: _op.neg}
_ALLOWED_CALLS = {"min": min, "max": max, "round": round, "abs": abs}


def safe_eval(expr: str, ctx: dict[str, Any]) -> Any:
    """Evaluate an arithmetic expression against ctx. Raises on unsafe nodes."""
    if expr is None or expr == "":
        return None
    tree = ast.parse(str(expr), mode="eval")
    return _eval_node(tree.body, ctx)


def _eval_node(node, ctx):
    if isinstance(node, ast.Constant):
        return node.value
    if isinstance(node, ast.Name):
        if node.id in ctx:
            return ctx[node.id]
        if node.id in _ALLOWED_CALLS:
            return _ALLOWED_CALLS[node.id]
        raise ValueError(f"unknown name: {node.id}")
    if isinstance(node, ast.BinOp):
        op = _BIN_OPS.get(type(node.op))
        if not op:
            raise ValueError(f"disallowed op: {type(node.op).__name__}")
        return op(_eval_node(node.left, ctx), _eval_node(node.right, ctx))
    if isinstance(node, ast.UnaryOp):
        op = _UNARY_OPS.get(type(node.op))
        if not op:
            raise ValueError(f"disallowed unary op")
        return op(_eval_node(node.operand, ctx))
    if isinstance(node, ast.Call):
        fn = _eval_node(node.func, ctx)
        if fn not in _ALLOWED_CALLS.values():
            raise ValueError("disallowed call")
        args = [_eval_node(a, ctx) for a in node.args]
        return fn(*args)
    if isinstance(node, ast.IfExp):
        return _eval_node(node.body, ctx) if _eval_node(node.test, ctx) else _eval_node(node.orelse, ctx)
    if isinstance(node, ast.Compare):
        left = _eval_node(node.left, ctx)
        for op, comp in zip(node.ops, node.comparators):
            right = _eval_node(comp, ctx)
            cmp_fn = {
                ast.Eq: _op.eq, ast.NotEq: _op.ne,
                ast.Lt: _op.lt, ast.LtE: _op.le,
                ast.Gt: _op.gt, ast.GtE: _op.ge,
            }.get(type(op))
            if not cmp_fn or not cmp_fn(left, right):
                return False
            left = right
        return True
    raise ValueError(f"disallowed node: {type(node).__name__}")


# ── Template-string interpolation: {{pct(...)}} / {{var}} ────────
_TMPL_RE = re.compile(r"\{\{\s*([^}]+?)\s*\}\}")


def _pct(v: float | None, decimals: int = 0) -> str:
    if v is None:
        return ""
    fmt = f"{{:.{decimals}f}}%"
    return fmt.format(round(float(v) * 100, decimals))


def render_text(text: str | None, ctx: dict[str, Any]) -> str | None:
    if text is None:
        return None

    def repl(m):
        expr = m.group(1).strip()
        # Special: pct(x) / pct(x, n)
        if expr.startswith("pct(") and expr.endswith(")"):
            inner = expr[4:-1]
            parts = [p.strip() for p in inner.split(",")]
            val = safe_eval(parts[0], ctx)
            decimals = int(safe_eval(parts[1], ctx)) if len(parts) > 1 else 0
            return _pct(val, decimals)
        try:
            v = safe_eval(expr, ctx)
            if v is None:
                return ""
            if isinstance(v, float):
                return f"{v:,.2f}"
            return str(v)
        except Exception:
            return m.group(0)

    return _TMPL_RE.sub(repl, text)


# ── Openpyxl style helpers ──────────────────────────────────────
_thin = Side(style="thin")
_border_all = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_border_bottom = Border(bottom=_thin)

_font_bold = Font(name="Calibri", size=11, bold=True)
_font_bold_brown = Font(name="Calibri", size=11, bold=True, color="8B4513")
_font_big_bold = Font(name="Calibri", size=13, bold=True)

_fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_fill_light_blue = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
_fill_orange = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
_fill_green = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

_FILLS = {"yellow": _fill_yellow, "light_blue": _fill_light_blue, "orange": _fill_orange, "green": _fill_green}
_BORDERS = {"all": _border_all, "bottom": _border_bottom}

_align_right = Alignment(horizontal="right")
_align_center = Alignment(horizontal="center")
_align_left = Alignment(horizontal="left")

NUM_2DP = "#,##0.00"


def build_context(
    *,
    revenue: float,
    manual_inputs: dict,
    location_name: str,
    params: dict,
    share_basis: str,
) -> dict[str, Any]:
    """Assemble evaluator context from live inputs + template params."""
    share_rate = float(manual_inputs.get("location_share_rate", 0.40))
    electricity = float(manual_inputs.get("electricity_cost", 0))
    internet = float(manual_inputs.get("internet_cost", 0))
    etax = float(manual_inputs.get("etax", 0))
    vat_rate = float(params.get("vat_rate", manual_inputs.get("vat_rate", 0.07)))
    tx_rate = float(params.get("tx_rate", manual_inputs.get("transaction_fee_rate", 0.0365)))
    wht_rate = float(params.get("wht_rate", 0.03))
    transfer_fee = float(params.get("transfer_fee", 30))
    evse_count = manual_inputs.get("evse_count") or 0

    sharge_rate = max(0.0, 1.0 - share_rate)

    ctx: dict[str, Any] = {
        "revenue": revenue,
        "electricity_cost": electricity,
        "internet_cost": internet,
        "etax": etax,
        "evse_count": evse_count,
        "location_name": location_name,
        "group_name": manual_inputs.get("group_name") or "",
        "share_rate": share_rate,
        "sharge_rate": sharge_rate,
        "vat_rate": vat_rate,
        "tx_rate": tx_rate,
        "wht_rate": wht_rate,
        "transfer_fee": transfer_fee,
    }

    # location_share — headline figure for "share" / "net_gp" rows.
    t = compute_totals(
        revenue, electricity, internet, etax,
        tx_fee_rate=tx_rate, share_rate=share_rate,
        share_basis=share_basis if share_basis in ("gp", "revenue") else "gp",
    )
    ctx["location_share"] = t["location_share"]
    ctx["remaining"] = t["remaining"]
    ctx["before_vat"] = t["before_vat"]
    ctx["vat_portion"] = t["vat_portion"]

    # Dealer-settlement derived helpers — always exposed so any template can
    # opt into them (the dealer_new_model template does; other revenue-basis
    # templates can ignore them).
    total_payment_to_sharge = (
        revenue * sharge_rate
        + etax + etax * vat_rate
        + internet + internet * vat_rate
    )
    wht_sharge_to_dealer = (revenue / (1 + vat_rate)) * wht_rate
    wht_dealer_to_sharge = (total_payment_to_sharge / (1 + vat_rate)) * wht_rate
    sharge_expense_net = total_payment_to_sharge - wht_dealer_to_sharge
    total_payment_to_dealer = revenue - sharge_expense_net - wht_sharge_to_dealer
    ctx.update({
        "total_payment_to_sharge": total_payment_to_sharge,
        "wht_sharge_to_dealer": wht_sharge_to_dealer,
        "wht_dealer_to_sharge": wht_dealer_to_sharge,
        "sharge_expense_net": sharge_expense_net,
        "total_payment_to_dealer": total_payment_to_dealer,
    })

    return ctx


def render_summary(ws, template: dict, ctx: dict[str, Any]) -> None:
    """Render `template.summary_layout` rows into worksheet `ws`.

    Kinds (affect styling + column mapping):
      * default           → 4-col layout [B=label, C=note, D=value]  (standard_gp/revenue_share)
      * header            → 3-col boxed header (row 2 "Revenue")
      * share             → light-blue highlighted share rows
      * net_gp            → Net GP big-bold row
      * dealer_header     → green header row (dealer template)
      * section           → section label only (e.g. "Expense")

    The layout's "kind" is the visual template — any template can reuse them.
    """
    layout = template.get("summary_layout") or []
    layout_style = template.get("layout_style") or "standard"

    if layout_style == "dealer":
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 38
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 16
    else:
        ws.column_dimensions["A"].width = 3
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 38
        ws.column_dimensions["D"].width = 18

    for row_cfg in layout:
        _render_row(ws, row_cfg, ctx, layout_style)


def _render_row(ws, row_cfg: dict, ctx: dict, layout_style: str) -> None:
    r = int(row_cfg.get("row", 0)) or None
    if r is None:
        return
    kind = row_cfg.get("kind") or "default"
    label = render_text(row_cfg.get("label"), ctx)
    note = render_text(row_cfg.get("note"), ctx)
    value_expr = row_cfg.get("value")
    try:
        value = safe_eval(value_expr, ctx) if value_expr else None
    except Exception:
        value = None

    fill = _FILLS.get(row_cfg.get("fill"))
    border = _BORDERS.get(row_cfg.get("border"))
    bold = bool(row_cfg.get("bold"))

    if layout_style == "dealer":
        _render_dealer_row(ws, r, kind, label, note, value, fill=fill, bold=bold)
    else:
        _render_std_row(ws, r, kind, label, note, value, fill=fill, bold=bold, border=border)


def _render_std_row(ws, r, kind, label, note, value, *, fill, bold, border):
    """Standard 4-col layout used by standard_gp / revenue_share."""
    if kind == "header":
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).border = _border_all
        ws.cell(row=r, column=2, value=label or "Revenue").font = _font_bold
        ws.cell(row=r, column=2).alignment = _align_center
        c = ws.cell(row=r, column=4, value=_round(value))
        c.font = _font_bold
        c.number_format = NUM_2DP
        c.alignment = _align_right
        return

    if kind == "net_gp":
        ws.cell(row=r, column=2, value=label or "Net GP").font = _font_big_bold
        ws.cell(row=r, column=2).alignment = _align_center
        ws.cell(row=r, column=3, value=note).font = _font_bold
        ws.cell(row=r, column=3).alignment = _align_center
        c = ws.cell(row=r, column=4, value=_round(value))
        c.font = _font_big_bold
        c.number_format = NUM_2DP
        c.alignment = _align_right
        c.border = Border(top=Side(style="double"), bottom=Side(style="double"))
        return

    if kind == "share":
        cb = ws.cell(row=r, column=2, value=label)
        cc = ws.cell(row=r, column=3, value=note)
        cd = ws.cell(row=r, column=4, value=_round(value))
        cd.number_format = NUM_2DP
        cd.alignment = _align_right
        cb.alignment = _align_right
        cc.alignment = _align_center
        for c in (cb, cc, cd):
            c.fill = _fill_light_blue
            c.border = _border_all
            c.font = _font_bold_brown
        return

    # Default 4-col row
    cb = ws.cell(row=r, column=2, value=label)
    if bold:
        cb.font = _font_bold
    cb.alignment = _align_right
    cc = ws.cell(row=r, column=3, value=note)
    cc.alignment = _align_center
    cd = ws.cell(row=r, column=4, value=_round(value))
    cd.number_format = NUM_2DP
    cd.alignment = _align_right
    if bold:
        cd.font = _font_bold
    if fill:
        cd.fill = fill
    if border:
        cd.border = border


def _render_dealer_row(ws, r, kind, label, note, value, *, fill, bold):
    """Dealer-settlement 5-col layout (B=label, C=note, D=THB, E=value)."""
    if kind == "dealer_header":
        cb = ws.cell(row=r, column=2, value=label)
        cb.font = _font_bold
        cb.fill = _fill_green
        cc = ws.cell(row=r, column=3, value=note)
        cc.fill = _fill_green
        cc.alignment = _align_left
        cd = ws.cell(row=r, column=4, value="THB")
        cd.fill = _fill_green
        ce = ws.cell(row=r, column=5, value=_round(value))
        ce.font = _font_bold
        ce.fill = _fill_green
        ce.number_format = NUM_2DP
        ce.alignment = _align_right
        return

    if kind == "section":
        ws.cell(row=r, column=2, value=label).font = _font_bold
        return

    # Default dealer row
    cb = ws.cell(row=r, column=2, value=label)
    cc = ws.cell(row=r, column=3, value=note)
    cc.alignment = _align_left
    cd = ws.cell(row=r, column=4, value="THB" if value is not None else "")
    ce = ws.cell(row=r, column=5, value=_round(value))
    ce.number_format = NUM_2DP
    ce.alignment = _align_right
    if bold:
        cb.font = _font_bold
        ce.font = _font_bold
    if fill:
        for c in (cb, cc, cd, ce):
            c.fill = fill


def _round(v):
    if v is None:
        return None
    try:
        return round(float(v), 2)
    except Exception:
        return v
