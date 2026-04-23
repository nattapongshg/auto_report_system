"""Build branded Excel report with Summary sheet + Data sheet + optional bill image."""

import os
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter


# ── Styles ──
thin = Side(style="thin")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_bottom = Border(bottom=thin)

font_header = Font(name="Calibri", size=11, bold=True)
font_data = Font(name="Calibri", size=11)
font_bold = Font(name="Calibri", size=11, bold=True)
font_bold_brown = Font(name="Calibri", size=11, bold=True, color="8B4513")
font_big_bold = Font(name="Calibri", size=13, bold=True)

fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
fill_light_blue = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
fill_orange = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
fill_green = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
fill_light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
font_green = Font(name="Calibri", size=11, color="006100")
font_red = Font(name="Calibri", size=11, color="C00000")     # full refund
font_orange = Font(name="Calibri", size=11, color="C65911")  # partial refund

align_right = Alignment(horizontal="right")
align_center = Alignment(horizontal="center")
num_2dp = "#,##0.00"

REPORT_HEADERS = [
    "Invoice", "Reference Id", "Location", "Evse", "Start Date Time", "End Date Time",
    "Payment Created At", "Transaction Id", "Invoice Status", "Payment Amount",
    "Total Cost", "Total Discount", "Unit Price ", "Kwh",
    "Total Time (Hour)", "Total Overtime (Hour)", "Etax Number", None,
    "Privilege Name", "RFID Number", "organization_name",
]


def _map_payment_row(row: dict) -> list:
    """Map a payment-level row dict (from payment_rows) to Excel row.

    One invoice with multiple payments becomes multiple Excel rows:
      - Primary (is_invoice_primary=True): full fields populated
      - Secondary: only reference_id / location / session dates / payment
        amount / privilege name — mirrors the manual-team layout.
    """
    is_primary = bool(row.get("is_invoice_primary"))
    kwh = float(row.get("kwh") or 0)
    invoice_amount = float(row.get("invoice_amount") or 0)
    payment_amount = float(row.get("payment_amount") or 0)
    refund_amount = float(row.get("refund_amount") or 0)
    net_amount = float(row.get("net_amount") or 0)
    refund_type = (row.get("refund_type") or "").lower()
    revenue = float(row.get("_revenue") or 0)
    privilege = row.get("privilege_program_name") or ""
    discount_label = row.get("discount_label") or ""
    invoice_status = row.get("invoice_status") or ""

    # Payment Amount column = revenue contribution of this payment row.
    pay_col = revenue

    if is_primary:
        # Show unit_price from invoice_amount / kwh when available
        unit_price = invoice_amount / kwh if kwh > 0 and invoice_amount > 0 else None
        total_cost = invoice_amount
        total_discount = float(row.get("total_discount") or 0)
        return [
            row.get("invoice_id"),
            row.get("reference_id"),
            row.get("location_name"),
            row.get("evse_name"),
            _parse_dt(row.get("session_start_bkk")),
            _parse_dt(row.get("session_end_bkk")),
            _parse_dt(row.get("payment_created_bkk")),
            row.get("payment_transaction_id"),
            invoice_status,
            pay_col,
            total_cost,
            total_discount,
            unit_price, kwh,
            float(row.get("total_time") or 0),
            float(row.get("total_overtime") or 0),
            row.get("etax_number"),
            None,
            privilege,
            "",  # rfid
            "",  # organization
        ]

    # Secondary payment — minimal fields, mirrors reference layout.
    return [
        row.get("invoice_id"),
        row.get("reference_id"),
        row.get("location_name"),
        row.get("evse_name"),
        _parse_dt(row.get("session_start_bkk")),
        _parse_dt(row.get("session_end_bkk")),
        _parse_dt(row.get("payment_created_bkk")),
        None, None,        # Transaction Id, Invoice Status — blank
        pay_col,            # Payment Amount
        None, None,        # Total Cost, Total Discount
        None, None,        # Unit Price, Kwh
        None, None,        # Total Time, Total Overtime
        None, None,        # Etax Number, blank
        privilege,
        "",
        "",
    ]


def _parse_dt(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(str(s).replace("Z", "").replace("+00:00", ""))
    except Exception:
        return s


def _map_row(row_dict: dict) -> list:
    """Map processed row dict to Excel row.

    For credit privileges with share_rate:
      - Unit Price = share_rate
      - Total Discount = kwh * share_rate (recalculated)
      - Payment Amount = kwh * share_rate (= revenue)
    For percent privileges or no privilege:
      - Unit Price = invoice_amount / kwh
      - Total Discount = raw total_discount
      - Payment Amount = _revenue (calculated)
    """
    # Auto-route to payment-level mapper when row comes from payment_rows
    # (marker: has 'is_invoice_primary' key). One invoice → multiple rows.
    if "is_invoice_primary" in row_dict:
        return _map_payment_row(row_dict)

    kwh = float(row_dict.get("kwh") or 0)
    invoice_amount = float(row_dict.get("invoice_amount") or 0)
    payment_amount = float(row_dict.get("payment_amount") or 0)
    raw_discount = float(row_dict.get("total_discount") or 0)
    raw_refund = float(row_dict.get("total_refund") or 0)
    revenue = float(row_dict.get("_revenue") or 0)
    share_rate = row_dict.get("_share_rate")
    priv_type = row_dict.get("_privilege_type")

    invoice_status = row_dict.get("invoice_status") or ""

    # "Payment-method change" — cash paid then swapped to credit → refund
    # cancels the cash leg. Display as a normal invoice (no refund shown).
    dstatus = (row_dict.get("discount_status") or "").strip().lower()
    is_method_change = (
        raw_refund > 0 and abs(raw_refund - payment_amount) < 0.01
        and dstatus == "paid" and invoice_status != "refunded"
    )
    refund_display = 0.0 if is_method_change else raw_refund

    # billed_to_organization: use invoice_amount directly (org pays at their rate)
    if invoice_status == "billed_to_organization":
        unit_price = invoice_amount / kwh if kwh > 0 and invoice_amount > 0 else None
        return [
            row_dict.get("invoice_id"), row_dict.get("reference_id"),
            row_dict.get("location_name"), row_dict.get("evse_name"),
            _parse_dt(row_dict.get("session_start_bkk")), _parse_dt(row_dict.get("session_end_bkk")),
            _parse_dt(row_dict.get("paid_date_bkk")), row_dict.get("payment_transaction_id"),
            invoice_status, invoice_amount, invoice_amount, 0,
            unit_price, kwh,
            float(row_dict.get("total_time") or 0), float(row_dict.get("total_overtime") or 0),
            row_dict.get("etax_number"), None,
            row_dict.get("_display_name"),
            row_dict.get("rfid_number") or "",
            row_dict.get("organization_name") or "",
        ]

    # Determine if this is a credit row that needs recalculation
    is_credit_row = (priv_type == "credit") or \
                    (priv_type == "mixed" and payment_amount == 0) or \
                    (priv_type is None and payment_amount == 0 and raw_discount > 0)

    if is_credit_row and share_rate and share_rate > 0:
        # Recalculate with share_rate
        unit_price = share_rate
        total_discount = kwh * share_rate
        pay_amount = total_discount  # revenue = recalculated discount
    elif is_credit_row:
        # No share_rate: use raw total_discount as-is
        unit_price = raw_discount / kwh if kwh > 0 else None
        total_discount = raw_discount
        pay_amount = raw_discount
    else:
        # Normal / percent: use original values
        unit_price = invoice_amount / kwh if kwh > 0 and invoice_amount > 0 else None
        total_discount = raw_discount
        pay_amount = revenue

    return [
        row_dict.get("invoice_id"),
        row_dict.get("reference_id"),
        row_dict.get("location_name"),
        row_dict.get("evse_name"),
        _parse_dt(row_dict.get("session_start_bkk")),
        _parse_dt(row_dict.get("session_end_bkk")),
        _parse_dt(row_dict.get("paid_date_bkk")),
        row_dict.get("payment_transaction_id"),
        row_dict.get("invoice_status"),
        pay_amount,
        invoice_amount,
        total_discount,
        unit_price,
        kwh,
        float(row_dict.get("total_time") or 0),
        float(row_dict.get("total_overtime") or 0),
        row_dict.get("etax_number"),
        None,
        row_dict.get("_display_name"),
        row_dict.get("rfid_number") or "",
        row_dict.get("organization_name") or "",
    ]


def _write_summary_row(ws, row, label, desc, value, label_font=font_bold,
                       val_fill=None, val_font=None, blue_row=False, border=None):
    c_b = ws.cell(row=row, column=2, value=label)
    c_b.font = label_font
    c_b.alignment = align_right
    c_c = ws.cell(row=row, column=3, value=desc)
    c_c.alignment = align_center
    c_d = ws.cell(row=row, column=4, value=round(value, 2) if value is not None else None)
    c_d.number_format = num_2dp
    c_d.alignment = align_right
    if val_fill:
        c_d.fill = val_fill
    if val_font:
        c_d.font = val_font
    if blue_row:
        for c in (c_b, c_c, c_d):
            c.fill = fill_light_blue
            c.border = border_all
            c.font = font_bold_brown
    if border:
        c_d.border = border


def _build_new_model_summary(
    ws,
    location_name: str,
    revenue: float,
    *,
    share_rate: float,
    internet_cost: float,
    etax: float,
    electricity_cost: float,
    vat_rate: float,
    evse_count: int | None = None,
) -> None:
    """Showroom New Model dealer-settlement layout.

    Rows:
      [top] Total Revenue (row 2)

      Expense block (rows 5-12):
        Net Revenue SHARGE   — revenue × share_rate (e.g. 10%)
        VAT                   — NetRev - (NetRev/1.07)
        Revenue SHARGE        — NetRev / 1.07      (yellow)
        Etax service fee      — etax baseline (per-doc × 1)
        Etax service fee Vat  — etax × 7%
        Internet              — (evse_count or 2) × 598
        Internet VAT          — internet × 7%
        Total Payment to SHARGE (orange)

      Dealer block (rows 15-20):
        Revenue Dealer        — same as Total Revenue (green)
        WHT Sharge→Dealer     — (Revenue/1.07) × 3%
        Sharge Fee            — Total Payment to SHARGE
        WHT Dealer→Sharge     — (SFee/1.07) × 3%
        ค่าใช้จ่าย Sharge หัก Dealer
        Total Payment to Dealer (yellow)

      Final block (rows 23-25):
        Revenue Dealer (Include VAT)
        Electricity Usage (Include VAT)
        Net Profit Dealer (orange)
    """
    from openpyxl.styles import Alignment

    # Column widths for this layout
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 16

    # Formulas
    # Dealer reports: location_share_rate = the *dealer's* cut (e.g. 0.90),
    # so Sharge's top-line cut is the inverse (0.10).
    sharge_rate = max(0.0, 1.0 - share_rate)
    net_rev_sharge = revenue * sharge_rate               # Sharge's top-line (incl VAT)
    rev_sharge_before_vat = net_rev_sharge / (1 + vat_rate)
    vat_on_rev_sharge = net_rev_sharge - rev_sharge_before_vat
    etax_vat = etax * vat_rate
    # Internet = operator-edited value on Locations page. Sim count shown as
    # informational label derived from evse count.
    internet = float(internet_cost)
    n_sims_label = f"{evse_count} Sims" if evse_count and evse_count > 0 else ""
    internet_vat_amt = internet * vat_rate
    total_payment_to_sharge = net_rev_sharge + etax + etax_vat + internet + internet_vat_amt

    wht_rate = 0.03
    wht_sharge_to_dealer = (revenue / (1 + vat_rate)) * wht_rate
    sharge_fee = total_payment_to_sharge
    wht_dealer_to_sharge = (sharge_fee / (1 + vat_rate)) * wht_rate
    sharge_expense_net = sharge_fee - wht_dealer_to_sharge
    total_payment_to_dealer = revenue - sharge_expense_net - wht_sharge_to_dealer
    net_profit_dealer = total_payment_to_dealer - electricity_cost

    # Header row labels
    ws.cell(row=2, column=2, value="Total Revenue").font = font_bold
    ws.cell(row=2, column=3, value="Include VAT").alignment = Alignment(horizontal="left")
    ws.cell(row=2, column=4, value="THB")
    ws.cell(row=2, column=5, value=round(revenue, 2)).font = font_bold
    ws.cell(row=2, column=5).number_format = num_2dp
    ws.cell(row=2, column=5).alignment = align_right

    def _line(r, label, note, value, *, fill=None, bold=False, border=None):
        cb = ws.cell(row=r, column=2, value=label)
        cc = ws.cell(row=r, column=3, value=note)
        cd = ws.cell(row=r, column=4, value="THB" if value is not None else "")
        ce = ws.cell(row=r, column=5, value=round(value, 2) if value is not None else None)
        ce.number_format = num_2dp
        ce.alignment = align_right
        cc.alignment = Alignment(horizontal="left")
        if bold:
            cb.font = font_bold
            ce.font = font_bold
        if fill:
            for c in (cb, cc, cd, ce):
                c.fill = fill
        if border:
            for c in (cb, cc, cd, ce):
                c.border = border

    # Expense section header
    ws.cell(row=4, column=2, value="Expense").font = font_bold

    _line(5, "Net Revenue SHARGE",  f"{round(sharge_rate*100)}% Top line revenue", net_rev_sharge)
    _line(6, "VAT",                 f"{int(vat_rate*100)}% of Revenue SHARGE", vat_on_rev_sharge)
    _line(7, "Revenue SHARGE",      "Before VAT", rev_sharge_before_vat, fill=fill_yellow, bold=True)
    _line(8, "Etax sertvice fee",   None, etax)
    _line(9, "Etax service fee (Vat)", f"{int(vat_rate*100)}% of etax service fee", etax_vat)
    _line(10, "Internet",           n_sims_label, internet)
    _line(11, "Internet (Vat)",     f"Vat {int(vat_rate*100)}% for sims", internet_vat_amt)
    _line(12, "Total Payment to SHARGE", None, total_payment_to_sharge, fill=fill_orange, bold=True)

    # Dealer settlement
    ws.cell(row=14, column=2, value="Revenue Dealer (Dealer invoice SHARGE) Include VAT").font = font_bold
    ws.cell(row=14, column=2).fill = fill_green
    ws.cell(row=14, column=3).fill = fill_green
    ws.cell(row=14, column=4, value="THB").fill = fill_green
    ws.cell(row=14, column=5, value=round(revenue, 2)).font = font_bold
    ws.cell(row=14, column=5).fill = fill_green
    ws.cell(row=14, column=5).number_format = num_2dp
    ws.cell(row=14, column=5).alignment = align_right

    _line(15, "WHT(Sharge=>Dealer)", None, wht_sharge_to_dealer)
    _line(16, "Sharge Fee (Include VAT)", None, sharge_fee)
    _line(17, "WHT(Dealer=>Sharge)", None, wht_dealer_to_sharge)
    _line(18, "ค่าใช้จ่าย Sharge หัก Dealer", None, sharge_expense_net)
    _line(19, "Total Payment to Dealer", None, total_payment_to_dealer, fill=fill_yellow, bold=True)

    # Net profit
    _line(22, "Revenue Dealer", "Include VAT", revenue)
    _line(23, "Electricity Usage (Include VAT)", None, electricity_cost)
    _line(24, "Net Profit Dealer", None, net_profit_dealer, fill=fill_orange, bold=True)


def build_report(
    rows: list[dict],
    location_name: str,
    manual_inputs: dict,
    bill_image_path: str | None = None,
) -> BytesIO:
    """Build complete Excel report and return as BytesIO."""

    # ── Extract manual inputs ──
    electricity_cost = float(manual_inputs.get("electricity_cost", 0))
    internet_cost = float(manual_inputs.get("internet_cost", 0))
    etax = float(manual_inputs.get("etax", 0))
    tx_fee_rate = float(manual_inputs.get("transaction_fee_rate", 0.0365))
    vat_rate = float(manual_inputs.get("vat_rate", 0.07))
    location_share_rate = float(manual_inputs.get("location_share_rate", 0.40))
    share_basis = manual_inputs.get("share_basis", "gp")
    group_name = manual_inputs.get("group_name", "")
    date_start = manual_inputs.get("date_start", "")
    date_end = manual_inputs.get("date_end", "")

    # ── Calculate revenue using same logic as _map_row ──
    revenue = 0
    for r in rows:
        mapped = _map_row(r)
        revenue += float(mapped[9] or 0)  # column index 9 = Payment Amount
    from app.engine.share_calc import compute_totals
    t = compute_totals(
        revenue, electricity_cost, internet_cost, etax,
        tx_fee_rate=tx_fee_rate, share_rate=location_share_rate,
        share_basis=share_basis,
    )
    tx_fee = t["tx_fee"]; vat_on_fee = t["vat_on_fee"]; total_fee = t["total_fee"]
    transfer_fee = t.get("transfer", 0)
    internet_incl_vat = t["internet_incl_vat"]; etax_incl_vat = t["etax_incl_vat"]
    remaining = t["remaining"]; location_share = t["location_share"]
    vat_portion = t["vat_portion"]; before_vat = t["before_vat"]

    # Sort by paid_date
    rows.sort(key=lambda r: r.get("paid_date_bkk") or "")

    wb = Workbook()

    # ═══════════════════════════════════
    # Summary Sheet
    # ═══════════════════════════════════
    ws = wb.active
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 18

    # Revenue
    for col in (2, 3, 4):
        ws.cell(row=2, column=col).border = border_all
    ws.cell(row=2, column=2, value="Revenue").font = font_bold
    ws.cell(row=2, column=2).alignment = align_center
    ws.cell(row=2, column=4, value=round(revenue, 2)).font = font_bold
    ws.cell(row=2, column=4).number_format = num_2dp
    ws.cell(row=2, column=4).alignment = align_right

    # Template-driven path: if the caller resolved a report_layout_template,
    # delegate summary rendering to the template engine. Data sheet + bill
    # image continue in the shared path below.
    template = manual_inputs.get("report_template")
    rendered_by_template = False
    if template:
        from app.engine.template_engine import build_context, render_summary
        ctx = build_context(
            revenue=revenue,
            manual_inputs=manual_inputs,
            location_name=location_name,
            params=template.get("params") or {},
            share_basis=template.get("share_basis") or "gp",
        )
        # Clear the pre-drawn row-2 "Revenue" header — template owns the full sheet.
        for col in (2, 3, 4, 5):
            ws.cell(row=2, column=col).value = None
            ws.cell(row=2, column=col).border = Border()
        render_summary(ws, template, ctx)
        rendered_by_template = True

    # ── Legacy fallback (used when caller didn't resolve a template) ──
    is_new_model = (not rendered_by_template) and group_name == "showroom_new_model"
    if rendered_by_template:
        pass
    elif is_new_model:
        _build_new_model_summary(
            ws, location_name, revenue,
            share_rate=location_share_rate,
            internet_cost=internet_cost,
            etax=etax,
            electricity_cost=electricity_cost,
            vat_rate=vat_rate,
            evse_count=manual_inputs.get("evse_count"),
        )
    elif share_basis != 'revenue':
        _write_summary_row(ws, 4, "Transaction Fee", f"({tx_fee_rate*100:.2f}% of Revenue)", tx_fee)
        _write_summary_row(ws, 5, "VAT", f"({vat_rate*100:.0f}% of Transaction Fee)", vat_on_fee)
        _write_summary_row(ws, 6, "Transfer", None, transfer_fee)
        _write_summary_row(ws, 7, "Total Fee", None, total_fee, border=border_bottom)

        _write_summary_row(ws, 9, "Electricity Cost", None, electricity_cost,
                           val_fill=fill_yellow, val_font=font_bold)
        _write_summary_row(ws, 10, "Internet Cost", None, internet_cost)
        _write_summary_row(ws, 11, None, "Vat 7%", internet_incl_vat)
        _write_summary_row(ws, 12, "Etax", None, etax)
        _write_summary_row(ws, 13, "Etax (Include Vat)", "Vat 7%", etax_incl_vat)
        _write_summary_row(ws, 14, "คงเหลือ", None, remaining,
                           val_font=font_bold, border=border_bottom)

    if not is_new_model and not rendered_by_template:
        share_note = (
            f"({int(location_share_rate*100)}% of Revenue)" if share_basis == 'revenue'
            else f"({int(location_share_rate*100)}% of Gross Profit VAT Incl.)"
        )
        _write_summary_row(ws, 16, location_name, share_note, location_share, blue_row=True)
        _write_summary_row(ws, 17, "VAT", "(7% of Cash In)", vat_portion, blue_row=True)
        ws.cell(row=18, column=3, value="(Before VAT)").alignment = align_center
        ws.cell(row=18, column=4, value=round(before_vat, 2)).number_format = num_2dp
        ws.cell(row=18, column=4).alignment = align_right
        for c in (2, 3, 4):
            ws.cell(row=18, column=c).fill = fill_light_blue
            ws.cell(row=18, column=c).border = border_all
            ws.cell(row=18, column=c).font = font_bold_brown

        # Net GP
        ws.cell(row=20, column=2, value="Net GP").font = font_big_bold
        ws.cell(row=20, column=2).alignment = align_center
        ws.cell(row=20, column=3, value="(VAT Included)").font = font_bold
        ws.cell(row=20, column=3).alignment = align_center
        ws.cell(row=20, column=4, value=round(location_share, 2)).font = font_big_bold
        ws.cell(row=20, column=4).number_format = num_2dp
        ws.cell(row=20, column=4).alignment = align_right
        ws.cell(row=20, column=4).border = Border(top=Side(style="double"), bottom=Side(style="double"))

    # Bill image
    if bill_image_path and os.path.exists(bill_image_path):
        img = XlImage(bill_image_path)
        max_w = 500
        ratio = max_w / img.width if img.width > max_w else 1
        img.width = int(img.width * ratio)
        img.height = int(img.height * ratio)
        ws.add_image(img, "F2")

    # ═══════════════════════════════════
    # Data Sheet
    # ═══════════════════════════════════
    month_label = date_start[:7].replace("-", ".") if date_start else "data"
    ws_data = wb.create_sheet(month_label)

    for col_idx, h in enumerate(REPORT_HEADERS, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=h)
        cell.font = font_header

    # Parse report month from date_start for cross-month detection
    report_month = None
    if date_start:
        try:
            report_month = datetime.fromisoformat(date_start).month
        except Exception:
            pass

    for row_num, row_dict in enumerate(rows, 2):
        mapped = _map_row(row_dict)
        status = (row_dict.get("invoice_status") or "").lower()
        refund = float(row_dict.get("total_refund") or 0)
        payment = float(row_dict.get("payment_amount") or 0)
        dstatus = (row_dict.get("discount_status") or "").strip().lower()
        # Payment-method change: refund equals payment AND credit was paid —
        # not a real refund, just a swap from cash to credit. Keep normal styling.
        method_change = (
            refund > 0 and abs(refund - payment) < 0.01
            and dstatus == "paid" and status != "refunded"
        )
        if method_change:
            row_font = font_data
        elif status == "refunded":
            row_font = font_red
        elif refund > 0:
            row_font = font_orange
        else:
            row_font = font_data
        refund_flag = row_font is not font_data
        for col_idx, val in enumerate(mapped, 1):
            cell = ws_data.cell(row=row_num, column=col_idx, value=val)
            cell.font = row_font
            if col_idx in (5, 6, 7) and isinstance(val, datetime):
                cell.number_format = "m/d/yy h:mm"
                # Highlight Start/End dates yellow if session is from different month
                if col_idx in (5, 6) and report_month and val.month != report_month:
                    cell.fill = fill_yellow
            elif col_idx == 10:
                cell.number_format = "0.00"
                cell.fill = fill_light_green
                # Keep refund color on the otherwise-green payment cell
                cell.font = row_font if refund_flag else font_green
            elif col_idx in (11, 12, 13, 14, 15, 16):
                cell.number_format = "0.00"

    # Auto-width
    for col_idx in range(1, len(REPORT_HEADERS) + 1):
        max_len = len(str(REPORT_HEADERS[col_idx - 1] or ""))
        for r in range(2, min(20, ws_data.max_row + 1)):
            val = ws_data.cell(row=r, column=col_idx).value
            max_len = max(max_len, min(len(str(val or "")), 40))
        ws_data.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # ── Write to BytesIO ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
