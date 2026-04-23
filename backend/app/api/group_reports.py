"""Group reports: consolidate many locations into one Excel + one email."""

import logging
import os
import json
from datetime import datetime, timezone

logger = logging.getLogger(__name__)

from fastapi import APIRouter, BackgroundTasks, HTTPException
from pydantic import BaseModel

from app.supabase_client import supabase
from app.engine.privilege_calc import process_rows, refresh_cache
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.engine.excel_builder import build_report
from app.db.raw_rows import load_snapshot_rows
from app.engine.email_service import send_report_email
from app.engine.share_calc import compute_totals

router = APIRouter(prefix="/group-reports", tags=["group-reports"])

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "output")

TX_FEE_RATE = 0.0365
VAT_RATE = 0.07
LOCATION_SHARE_RATE = 0.40
DEFAULT_INTERNET_COST = 598
DEFAULT_ETAX = 0


class GroupLocationInput(BaseModel):
    location_id: str
    electricity_cost: float = 0
    internet_cost: float = DEFAULT_INTERNET_COST
    etax: float = 0


class GroupSendRequest(BaseModel):
    group_name: str
    # Per-location inputs: aggregated server-side for cost calculation
    location_inputs: list[GroupLocationInput] = []
    share_rate: float | None = None           # override; default = mode across locations
    tx_fee_rate: float | None = None
    bill_image_urls: list[str] | None = None
    email_recipients: list[str] | None = None
    skip_email: bool = False


@router.get("/groups")
async def list_groups():
    """List all distinct group_names used across active locations."""
    locs = await supabase.select(
        "locations",
        "select=group_name&is_active=eq.true"
    )
    agg: dict[str, int] = {}
    for l in locs:
        g = l.get("group_name")
        if not g:
            continue
        agg[g] = agg.get(g, 0) + 1
    items = [{"group_name": g, "location_count": c} for g, c in sorted(agg.items())]
    return {"items": items, "total": len(items)}


@router.get("/groups/{group_name}/locations")
async def list_group_locations(group_name: str, snapshot_id: str | None = None):
    """Locations in a group with default cost values for the input table.

    If snapshot_id is provided, merges per-snapshot etax (computed from etax_number
    count during init) and last-saved electricity_cost from monthly_location_inputs.
    """
    locs = await supabase.select(
        "locations",
        f"select=id,name,station_code,electricity_cost,internet_cost,etax"
        f"&group_name=eq.{group_name}&is_active=eq.true&order=name.asc"
    )

    if snapshot_id and locs:
        # Fetch ALL entries for this snapshot then match by location_id —
        # avoids URL length issues with in.(...) on large groups.
        entries = await supabase.select(
            "monthly_location_inputs",
            f"snapshot_id=eq.{snapshot_id}"
            f"&select=location_id,electricity_cost,internet_cost,etax,preview_rows"
        )
        by_loc = {e["location_id"]: e for e in entries}
        for l in locs:
            e = by_loc.get(l["id"])
            if e:
                if e.get("etax") is not None:
                    l["etax"] = e["etax"]
                if e.get("internet_cost"):
                    l["internet_cost"] = e["internet_cost"]
                if e.get("electricity_cost") and float(e["electricity_cost"]) > 0:
                    l["electricity_cost"] = e["electricity_cost"]
                l["preview_rows"] = e.get("preview_rows")

    return {"items": locs, "total": len(locs)}


@router.get("/{snapshot_id}/preview/{group_name}")
async def preview_group(snapshot_id: str, group_name: str):
    """Quick aggregate preview for a group (rows, kwh, revenue) against a snapshot."""
    snap = await supabase.select("monthly_snapshots", f"id=eq.{snapshot_id}", single=True)
    if not snap or snap.get("status") != "completed":
        raise HTTPException(400, "Snapshot not ready")

    locs = await supabase.select(
        "locations",
        f"select=id,name&group_name=eq.{group_name}&is_active=eq.true"
    )
    loc_names = {l["name"] for l in locs}

    raw_rows, cols = await load_snapshot_rows(snap["id"])
    if not raw_rows:
        raise HTTPException(400, "Snapshot has no rows")

    await refresh_cache()
    processed = await process_rows(raw_rows, cols, None)
    group_rows = [r for r in processed if r.get("location_name") in loc_names]
    revenue = sum(float(r.get("_revenue") or 0) for r in group_rows)
    kwh = sum(float(r.get("kwh") or 0) for r in group_rows)

    return {
        "group_name": group_name,
        "location_count": len(locs),
        "rows": len(group_rows),
        "kwh": round(kwh, 2),
        "revenue": round(revenue, 2),
    }


@router.post("/{snapshot_id}/send")
async def send_group_report(snapshot_id: str, payload: GroupSendRequest,
                            background_tasks: BackgroundTasks):
    """Send consolidated group report: one Excel with aggregated summary + all rows."""
    snap = await supabase.select("monthly_snapshots", f"id=eq.{snapshot_id}", single=True)
    if not snap or snap.get("status") != "completed":
        raise HTTPException(400, "Snapshot not ready")

    # Upsert a tracking row
    existing = await supabase.select(
        "group_report_inputs",
        f"snapshot_id=eq.{snapshot_id}&group_name=eq.{payload.group_name}&limit=1"
    )
    total_elec = sum(i.electricity_cost for i in payload.location_inputs)
    total_internet = sum(i.internet_cost for i in payload.location_inputs)
    total_etax = sum(i.etax for i in payload.location_inputs)

    entry_data = {
        "snapshot_id": snapshot_id,
        "group_name": payload.group_name,
        "year_month": snap["year_month"],
        "electricity_cost": total_elec,
        "internet_cost": total_internet,
        "etax": total_etax,
        "bill_image_urls": payload.bill_image_urls or [],
        "status": "generating",
        "submitted_at": datetime.now(timezone.utc).isoformat(),
    }
    if existing:
        entry = await supabase.update(
            "group_report_inputs", f"id=eq.{existing[0]['id']}", entry_data
        )
    else:
        entry = await supabase.insert("group_report_inputs", entry_data)

    background_tasks.add_task(_generate_group_task, snap, payload, entry["id"])
    return {"processing": True, "entry_id": entry["id"]}


async def _generate_group_task(snap: dict, payload: GroupSendRequest, entry_id: str):
    try:
        snapshot_id = snap["id"]
        year_month = snap["year_month"]
        group = payload.group_name

        # 1. Fetch locations in group
        locs = await supabase.select(
            "locations",
            f"select=id,name,station_code,location_share_rate,transaction_fee_rate,share_basis,email_recipients,group_name"
            f"&group_name=eq.{group}&is_active=eq.true"
        )
        if not locs:
            raise RuntimeError(f"No active locations in group {group}")
        loc_names = {l["name"] for l in locs}

        # 2. Load snapshot rows from DB and filter to this group
        raw_rows, cols = await load_snapshot_rows(snap["id"])

        await refresh_cache()
        all_processed = await process_rows(raw_rows, cols, None)
        group_rows = [r for r in all_processed if r.get("location_name") in loc_names]

        if not group_rows:
            raise RuntimeError(f"No rows found for group {group} in {year_month}")

        # 3. Determine share_rate / tx_fee_rate — use mode across locations or override
        from collections import Counter
        share_rate = payload.share_rate
        if share_rate is None:
            rates = [float(l.get("location_share_rate") or LOCATION_SHARE_RATE) for l in locs]
            share_rate = Counter(rates).most_common(1)[0][0]

        tx_fee_rate = payload.tx_fee_rate
        if tx_fee_rate is None:
            rates = [float(l.get("transaction_fee_rate") or TX_FEE_RATE) for l in locs]
            tx_fee_rate = Counter(rates).most_common(1)[0][0]

        # Group share_basis = mode across member locations
        bases = [(l.get("share_basis") or 'gp') for l in locs]
        share_basis = Counter(bases).most_common(1)[0][0]

        # 4. Pick bill image (first uploaded one; embed inline)
        bill_path = None
        if payload.bill_image_urls:
            first = payload.bill_image_urls[0]
            local = os.path.join(os.path.dirname(__file__), "..", "..", "uploads",
                                 os.path.basename(first))
            if os.path.exists(local):
                bill_path = local

        # 5. Aggregate per-location inputs
        total_elec = sum(i.electricity_cost for i in payload.location_inputs)
        total_internet = sum(i.internet_cost for i in payload.location_inputs)
        total_etax = sum(i.etax for i in payload.location_inputs)

        # Build consolidated Excel — treat group_name as the "location_name"
        manual_inputs = {
            "date_start": f"{year_month}-01",
            "date_end": f"{year_month}-28",
            "electricity_cost": float(total_elec),
            "internet_cost": float(total_internet),
            "etax": float(total_etax),
            "transaction_fee_rate": tx_fee_rate,
            "location_share_rate": share_rate,
            "share_basis": share_basis,
            "vat_rate": VAT_RATE,
        }
        excel_bytes = build_report(
            rows=group_rows,
            location_name=f"{group} (Group — {len(locs)} locations)",
            manual_inputs=manual_inputs,
            bill_image_path=bill_path,
        )

        # Augment workbook with a per-location breakdown sheet
        excel_bytes = _add_by_location_sheet(
            excel_bytes=excel_bytes,
            locs=locs,
            group_rows=group_rows,
            payload_inputs=payload.location_inputs,
            tx_fee_rate=tx_fee_rate,
            share_rate=share_rate,
        )

        safe = group.replace(" ", "_").replace("/", "_")[:60]
        filename = f"GROUP_{safe}_{year_month}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, filename)
        with open(output_path, "wb") as f:
            f.write(excel_bytes.read())
        file_size = os.path.getsize(output_path)

        # 6. Compute preview numbers for DB
        revenue = sum(float(r.get("_revenue") or 0) for r in group_rows)
        kwh = sum(float(r.get("kwh") or 0) for r in group_rows)
        t = compute_totals(
            revenue, float(total_elec), float(total_internet), float(total_etax),
            tx_fee_rate=tx_fee_rate, share_rate=share_rate, share_basis=share_basis,
        )
        remaining = t["remaining"]
        share = t["location_share"]

        # 7. Resolve recipients
        skip_email = payload.skip_email
        if payload.email_recipients is not None:
            recipients = payload.email_recipients
        else:
            # fallback: union of all locations' recipients
            seen: set[str] = set()
            recipients = []
            for l in locs:
                for e in (l.get("email_recipients") or []):
                    if e and e not in seen:
                        seen.add(e)
                        recipients.append(e)

        email_sent_at = None
        email_error = None
        if recipients and not skip_email:
            summary = {
                "revenue": revenue,
                "tx_fee_rate": tx_fee_rate,
                "vat_rate": VAT_RATE,
                "location_share_rate": share_rate,
                "share_basis": share_basis,
                "electricity_cost": total_elec,
                "internet_cost": total_internet,
                "etax": total_etax,
                "location_name": f"{group} (Group)",
                **t,
            }
            result = send_report_email(
                to=recipients,
                location_name=f"{group} (Group)",
                year_month=year_month,
                file_path=output_path,
                file_name=filename,
                summary=summary,
            )
            if result.get("status") == "sent":
                email_sent_at = datetime.now(timezone.utc).isoformat()
            else:
                email_error = result.get("error") or "email rejected"

        final_status = "failed" if (recipients and not skip_email and not email_sent_at) else "sent"

        await supabase.update("group_report_inputs", f"id=eq.{entry_id}", {
            "status": final_status,
            "file_name": filename,
            "file_path": output_path,
            "file_size_bytes": file_size,
            "preview_rows": len(group_rows),
            "preview_revenue": round(revenue, 2),
            "preview_kwh": round(kwh, 2),
            "preview_gp": round(remaining, 2),
            "preview_share": round(share, 2),
            "location_count": len(locs),
            "email_sent_at": email_sent_at,
            "email_error": email_error,
        })
        logger.info("%s: %s (%d rows, %d locations)", group, final_status, len(group_rows), len(locs))

    except Exception as e:
        await supabase.update("group_report_inputs", f"id=eq.{entry_id}", {
            "status": "failed",
            "email_error": str(e),
        })
        logger.exception("group report generation failed")


def _add_by_location_sheet(excel_bytes: BytesIO, locs: list[dict], group_rows: list[dict],
                           payload_inputs: list, tx_fee_rate: float, share_rate: float) -> BytesIO:
    """Append a 'By Location' sheet with per-location breakdown to the group Excel."""
    # Aggregate metrics per location
    loc_info = {l["id"]: l for l in locs}
    name_to_id = {l["name"]: l["id"] for l in locs}
    inputs_by_id = {i.location_id: i for i in payload_inputs}

    per_loc: dict[str, dict] = {}
    for r in group_rows:
        lname = r.get("location_name")
        lid = name_to_id.get(lname)
        if not lid:
            continue
        bucket = per_loc.setdefault(lid, {"rows": 0, "kwh": 0.0, "revenue": 0.0})
        bucket["rows"] += 1
        bucket["kwh"] += float(r.get("kwh") or 0)
        bucket["revenue"] += float(r.get("_revenue") or 0)

    # Load workbook
    excel_bytes.seek(0)
    wb = load_workbook(excel_bytes)
    if "By Location" in wb.sheetnames:
        del wb["By Location"]
    ws = wb.create_sheet("By Location", index=1)  # after Summary

    # Styles
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="121212", end_color="121212", fill_type="solid")
    font_total = Font(name="Calibri", size=11, bold=True)
    fill_total = PatternFill(start_color="FFF8C5", end_color="FFF8C5", fill_type="solid")
    align_right = Alignment(horizontal="right")
    align_center = Alignment(horizontal="center")
    num_2dp = "#,##0.00"
    num_int = "#,##0"

    share_pct = round(share_rate * 100, 2)
    tx_pct = round(tx_fee_rate * 100, 2)
    headers = [
        "#", "Location", "Station Code", "Rows", "kWh", "Revenue",
        f"Tx Fee ({tx_pct}%)", "Tx Fee VAT (7%)", "Total Fee",
        "Electricity",
        "Internet", "Internet Incl. VAT",
        "eTax", "eTax Incl. VAT",
        "Net GP", f"Location Share ({share_pct}%)",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
        c.border = border

    totals = {
        "rows": 0, "kwh": 0.0, "revenue": 0.0,
        "tx_fee": 0.0, "tx_vat": 0.0, "total_fee": 0.0,
        "elec": 0.0,
        "internet": 0.0, "internet_vat": 0.0,
        "etax": 0.0, "etax_vat": 0.0,
        "net_gp": 0.0, "share": 0.0,
    }

    row_idx = 2
    for i, loc in enumerate(locs, 1):
        lid = loc["id"]
        metrics = per_loc.get(lid, {"rows": 0, "kwh": 0.0, "revenue": 0.0})
        inp = inputs_by_id.get(lid)
        elec = float(inp.electricity_cost) if inp else 0.0
        internet = float(inp.internet_cost) if inp else 0.0
        etax = float(inp.etax) if inp else 0.0

        revenue = metrics["revenue"]
        tx_fee = revenue * tx_fee_rate
        vat_on_fee = tx_fee * 0.07
        total_fee = tx_fee + vat_on_fee
        internet_incl_vat = internet * 1.07
        etax_incl_vat = etax * 1.07
        net_gp = revenue - total_fee - elec - internet_incl_vat - etax_incl_vat
        loc_share = net_gp * share_rate

        values = [
            i, loc["name"], loc.get("station_code") or "",
            metrics["rows"], round(metrics["kwh"], 2),
            round(revenue, 2),
            round(tx_fee, 2), round(vat_on_fee, 2), round(total_fee, 2),
            round(elec, 2),
            round(internet, 2), round(internet_incl_vat, 2),
            round(etax, 2), round(etax_incl_vat, 2),
            round(net_gp, 2), round(loc_share, 2),
        ]
        for col, v in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.border = border
            if col >= 4:
                c.alignment = align_right
                c.number_format = num_int if col == 4 else num_2dp
        row_idx += 1

        totals["rows"] += metrics["rows"]
        totals["kwh"] += metrics["kwh"]
        totals["revenue"] += revenue
        totals["tx_fee"] += tx_fee
        totals["tx_vat"] += vat_on_fee
        totals["total_fee"] += total_fee
        totals["elec"] += elec
        totals["internet"] += internet
        totals["internet_vat"] += internet_incl_vat
        totals["etax"] += etax
        totals["etax_vat"] += etax_incl_vat
        totals["net_gp"] += net_gp
        totals["share"] += loc_share

    # Total row
    total_values = [
        "", "TOTAL", f"{len(locs)} locations",
        totals["rows"], round(totals["kwh"], 2),
        round(totals["revenue"], 2),
        round(totals["tx_fee"], 2), round(totals["tx_vat"], 2), round(totals["total_fee"], 2),
        round(totals["elec"], 2),
        round(totals["internet"], 2), round(totals["internet_vat"], 2),
        round(totals["etax"], 2), round(totals["etax_vat"], 2),
        round(totals["net_gp"], 2), round(totals["share"], 2),
    ]
    for col, v in enumerate(total_values, 1):
        c = ws.cell(row=row_idx, column=col, value=v)
        c.font = font_total
        c.fill = fill_total
        c.border = border
        if col >= 4:
            c.alignment = align_right
            c.number_format = num_int if col == 4 else num_2dp

    # Column widths
    widths = [5, 40, 14, 8, 12, 14, 12, 12, 12, 14, 12, 14, 10, 12, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@router.get("/{snapshot_id}/history")
async def list_group_reports(snapshot_id: str):
    """List group reports created for a snapshot."""
    items = await supabase.select(
        "group_report_inputs",
        f"snapshot_id=eq.{snapshot_id}&order=group_name.asc"
    )
    return {"items": items, "total": len(items)}
