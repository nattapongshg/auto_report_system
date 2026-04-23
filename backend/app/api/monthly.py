"""Monthly snapshot + batch run API."""

import csv
import io
import logging
import os
import json
import asyncio
from datetime import datetime, timezone
from typing import Any

from fastapi import APIRouter, File, Form, HTTPException, BackgroundTasks, UploadFile
from pydantic import BaseModel

logger = logging.getLogger(__name__)

from app.supabase_client import supabase
from app.engine.fetcher import fetch_date_range
from app.engine.privilege_calc import process_rows, refresh_cache
from app.engine.excel_builder import build_report
from app.db.raw_rows import delete_snapshot_rows, insert_snapshot_rows, load_snapshot_rows
from app.db.payment_rows import delete_payment_rows, insert_payment_rows

router = APIRouter(prefix="/monthly", tags=["monthly"])

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "output")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


class FetchRequest(BaseModel):
    year_month: str  # e.g. '2026-03'
    question_id: int = 1144


class BatchRunRequest(BaseModel):
    snapshot_id: str


# ─── Snapshots ───

@router.get("/snapshots")
async def list_snapshots():
    items = await supabase.select("monthly_snapshots", "select=*&order=year_month.desc")
    return {"items": items, "total": len(items)}


@router.post("/snapshots/fetch")
async def fetch_snapshot(req: FetchRequest, background_tasks: BackgroundTasks):
    """Start fetching raw data from Metabase for a given month."""
    # Check if already exists
    existing = await supabase.select(
        "monthly_snapshots",
        f"year_month=eq.{req.year_month}&question_id=eq.{req.question_id}&limit=1"
    )
    if existing and existing[0]["status"] == "completed":
        return existing[0]

    if existing and existing[0]["status"] == "fetching":
        raise HTTPException(400, "Already fetching this month")

    # Create or update snapshot record
    if existing:
        snapshot = await supabase.update(
            "monthly_snapshots", f"id=eq.{existing[0]['id']}",
            {"status": "fetching", "error_message": None}
        )
    else:
        snapshot = await supabase.insert("monthly_snapshots", {
            "year_month": req.year_month,
            "question_id": req.question_id,
            "status": "fetching",
        })

    # Run fetch in background
    background_tasks.add_task(_fetch_snapshot_task, snapshot["id"], req.year_month, req.question_id)
    return snapshot


# ─── Upload raw file (Metabase export) ───
# Manual alternative to the Metabase fetch for environments without VPN
# access: operator exports Q1144 from Metabase (xlsx or csv) on the 1st of
# the month and uploads it here. Rows feed the same `metabase_rows` table so
# the rest of the report pipeline is unchanged.

# Column names accepted from Metabase exports. Metabase sometimes emits
# display names (title-case with spaces) — we normalise here.
_HEADER_ALIASES = {
    "invoice id": "invoice_id",
    "invoice_status": "invoice_status",
    "invoice status": "invoice_status",
    "etax number": "etax_number",
    "reference id": "reference_id",
    "session start bkk": "session_start_bkk",
    "session end bkk": "session_end_bkk",
    "paid date bkk": "paid_date_bkk",
    "user email": "user_email",
    "location name": "location_name",
    "location code": "location_code",
    "evse name": "evse_name",
    "total time": "total_time",
    "total overtime": "total_overtime",
    "total overtime cost": "total_overtime_cost",
    "invoice amount": "invoice_amount",
    "total discount": "total_discount",
    "total refund": "total_refund",
    "price per kwh": "price_per_kwh",
    "payment amount": "payment_amount",
    "payment status": "payment_status",
    "payment provider": "payment_provider",
    "payment transaction id": "payment_transaction_id",
    "discount label": "discount_label",
    "privilege program name": "privilege_program_name",
    "discount provider": "discount_provider",
    "discount status": "discount_status",
}


def _normalize_header(h: str) -> str:
    if not h:
        return ""
    key = str(h).strip().lower()
    return _HEADER_ALIASES.get(key, key.replace(" ", "_"))


def _parse_raw_file(filename: str, data: bytes) -> tuple[list[list[Any]], list[str]]:
    """Parse xlsx or csv export into (rows, col_names)."""
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    if ext == "xlsx" or data[:2] == b"PK":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header_raw = list(next(it))
        headers = [_normalize_header(h) for h in header_raw]
        rows = [list(r) for r in it if any(v is not None and v != "" for v in r)]
        return rows, headers
    # CSV
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    header_raw = next(reader)
    headers = [_normalize_header(h) for h in header_raw]
    rows = [list(r) for r in reader if any(v for v in r)]
    return rows, headers


def _filter_to_month(rows: list[list[Any]], headers: list[str], year_month: str) -> tuple[list[list[Any]], int]:
    """Drop rows whose paid_date_bkk falls outside the target month. Returns (filtered, dropped_count)."""
    if "paid_date_bkk" not in headers:
        return rows, 0
    idx = headers.index("paid_date_bkk")
    kept = []
    for r in rows:
        v = r[idx]
        if v is None or v == "":
            continue
        s = str(v)[:7]  # YYYY-MM
        if s == year_month:
            kept.append(r)
    return kept, len(rows) - len(kept)


@router.post("/snapshots/upload")
async def upload_snapshot(
    year_month: str = Form(...),
    question_id: int = Form(1144),
    file: UploadFile = File(...),
) -> dict[str, Any]:
    """Upload a Metabase export (.xlsx or .csv) as the raw snapshot for a
    month. Operator exports Q1144 manually and uploads here on the 1st."""
    data = await file.read()
    try:
        rows, headers = _parse_raw_file(file.filename or "", data)
    except Exception as e:
        logger.exception("parse failed")
        raise HTTPException(400, f"parse failed: {e}")

    if not rows:
        raise HTTPException(400, "no rows in uploaded file")

    kept, dropped = _filter_to_month(rows, headers, year_month)
    if not kept:
        raise HTTPException(400, f"no rows for {year_month} in file (check paid_date_bkk column)")

    # Create/replace snapshot record, then bulk insert
    existing = await supabase.select(
        "monthly_snapshots",
        f"year_month=eq.{year_month}&question_id=eq.{question_id}&limit=1",
    )
    if existing:
        snapshot = await supabase.update(
            "monthly_snapshots", f"id=eq.{existing[0]['id']}",
            {"status": "fetching", "error_message": None},
        )
    else:
        snapshot = await supabase.insert("monthly_snapshots", {
            "year_month": year_month,
            "question_id": question_id,
            "status": "fetching",
        })

    try:
        await delete_snapshot_rows(snapshot["id"])
        inserted = await insert_snapshot_rows(snapshot["id"], kept, headers)
        await supabase.update(
            "monthly_snapshots", f"id=eq.{snapshot['id']}",
            {
                "status": "completed",
                "total_rows": inserted,
                "fetched_at": datetime.now(timezone.utc).isoformat(),
            },
        )
    except Exception as e:
        await supabase.update(
            "monthly_snapshots", f"id=eq.{snapshot['id']}",
            {"status": "failed", "error_message": str(e)},
        )
        raise HTTPException(500, f"insert failed: {e}")

    return {
        "snapshot_id": snapshot["id"],
        "year_month": year_month,
        "inserted": inserted,
        "dropped_out_of_month": dropped,
        "source_file": file.filename,
    }


async def _fetch_snapshot_task(snapshot_id: str, year_month: str, question_id: int):
    """Background task to fetch monthly data."""
    try:
        # Parse date range. Metabase's PaidDate_end is inclusive, so pass the
        # last day of the month — otherwise paid_date falling on the 1st of
        # the next month leaks into the snapshot.
        import calendar as _cal
        parts = year_month.split("-")
        year, month = int(parts[0]), int(parts[1])
        last_day = _cal.monthrange(year, month)[1]
        date_start = f"{year_month}-01"
        date_end = f"{year_month}-{last_day:02d}"

        logger.info("Fetching Q#%s %s to %s", question_id, date_start, date_end)

        rows, col_names = await fetch_date_range(
            question_id, date_start, date_end,
            on_progress=lambda msg: logger.info("snapshot: %s", msg)
        )

        # Q1145 is payment-centric (rows are payments, not invoices). Store in
        # a parallel table so the invoice-centric pipeline stays untouched
        # during dual-run verification.
        is_payment_centric = question_id == 1145

        # Safety: drop rows outside the target month.
        date_col = "payment_created_bkk" if is_payment_centric else "paid_date_bkk"
        if date_col in col_names:
            idx = col_names.index(date_col)
            before = len(rows)
            rows = [r for r in rows if not r[idx] or str(r[idx])[:7] == year_month]
            if before != len(rows):
                logger.info("dropped %d rows outside %s %s window", before - len(rows), year_month, date_col)

        if is_payment_centric:
            await delete_payment_rows(snapshot_id)
            inserted = await insert_payment_rows(snapshot_id, rows, col_names)
        else:
            await delete_snapshot_rows(snapshot_id)
            inserted = await insert_snapshot_rows(snapshot_id, rows, col_names)

        await supabase.update("monthly_snapshots", f"id=eq.{snapshot_id}", {
            "status": "completed",
            "total_rows": inserted,
            "file_path": None,  # rows now live in metabase_rows
            "fetched_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info("Snapshot done: %d rows stored in DB", inserted)

    except Exception as e:
        await supabase.update("monthly_snapshots", f"id=eq.{snapshot_id}", {
            "status": "failed",
            "error_message": str(e),
        })
        logger.exception("snapshot fetch failed")


# ─── Batch Runs ───

@router.get("/batch-runs")
async def list_batch_runs():
    items = await supabase.select("batch_runs", "select=*&order=created_at.desc")
    return {"items": items, "total": len(items)}


@router.get("/batch-runs/{batch_id}")
async def get_batch_run(batch_id: str):
    batch = await supabase.select("batch_runs", f"id=eq.{batch_id}", single=True)
    if not batch:
        raise HTTPException(404, "Batch run not found")
    items = await supabase.select("batch_run_items", f"batch_run_id=eq.{batch_id}&order=location_name.asc")
    batch["items"] = items
    return batch


@router.post("/batch-runs")
async def create_batch_run(req: BatchRunRequest, background_tasks: BackgroundTasks):
    """Start batch report generation for all enabled locations."""
    snapshot = await supabase.select("monthly_snapshots", f"id=eq.{req.snapshot_id}", single=True)
    if not snapshot:
        raise HTTPException(404, "Snapshot not found")
    if snapshot["status"] != "completed":
        raise HTTPException(400, "Snapshot not ready")

    # Get enabled locations
    locations = await supabase.select("locations", "is_report_enabled=eq.true&order=name.asc")
    if not locations:
        raise HTTPException(400, "No locations enabled for reports")

    # Create batch run
    batch = await supabase.insert("batch_runs", {
        "snapshot_id": req.snapshot_id,
        "year_month": snapshot["year_month"],
        "status": "pending",
        "total_locations": len(locations),
    })

    # Create items for each location
    for loc in locations:
        await supabase.insert("batch_run_items", {
            "batch_run_id": batch["id"],
            "location_id": loc["id"],
            "location_name": loc["name"],
            "status": "pending",
        })

    # Run in background
    background_tasks.add_task(_batch_run_task, batch["id"], snapshot)
    return batch


async def _batch_run_task(batch_id: str, snapshot: dict):
    """Background task to generate reports for all locations."""
    try:
        await supabase.update("batch_runs", f"id=eq.{batch_id}", {
            "status": "running",
            "started_at": datetime.now(timezone.utc).isoformat(),
        })

        # Load raw rows from DB
        raw_rows, col_names = await load_snapshot_rows(snapshot["id"])
        year_month = snapshot["year_month"]

        logger.info("Batch loaded %d rows for snapshot %s", len(raw_rows), snapshot["id"])

        # Refresh privilege cache
        await refresh_cache()

        # Get all batch items
        items = await supabase.select("batch_run_items", f"batch_run_id=eq.{batch_id}&order=location_name.asc")

        completed = 0
        failed = 0

        for item in items:
            loc_name = item["location_name"]
            try:
                await supabase.update("batch_run_items", f"id=eq.{item['id']}", {"status": "running"})

                # Get location config
                locs = await supabase.select("locations", f"id=eq.{item['location_id']}")
                loc_config = locs[0] if locs else {}

                # Process rows for this location
                processed = await process_rows(raw_rows, col_names, loc_name)

                if not processed:
                    await supabase.update("batch_run_items", f"id=eq.{item['id']}", {
                        "status": "completed", "row_count": 0, "revenue": 0,
                    })
                    completed += 1
                    continue

                # Build manual inputs from location config
                manual_inputs = {
                    "date_start": f"{year_month}-01",
                    "date_end": f"{year_month}-28",  # approximate
                    "electricity_cost": float(loc_config.get("electricity_cost") or 0),
                    "internet_cost": float(loc_config.get("internet_cost") or 0),
                    "etax": float(loc_config.get("etax") or 0),
                    "transaction_fee_rate": float(loc_config.get("transaction_fee_rate") or 0.0365),
                    "location_share_rate": float(loc_config.get("location_share_rate") or 0.40),
                    "share_basis": loc_config.get("share_basis") or "gp",
                }

                # Generate Excel
                excel_bytes = build_report(
                    rows=processed,
                    location_name=loc_name,
                    manual_inputs=manual_inputs,
                )

                safe_name = loc_name.replace(" ", "_").replace("/", "_")[:50]
                filename = f"{safe_name}_{year_month}.xlsx"
                output_path = os.path.join(OUTPUT_DIR, filename)

                with open(output_path, "wb") as f:
                    f.write(excel_bytes.read())

                revenue = sum(float(r.get("_revenue", 0)) for r in processed)

                # Send email if recipients configured
                email_recipients = loc_config.get("email_recipients") or []
                email_status = "completed"
                if email_recipients:
                    from app.engine.email_service import send_report_email
                    email_result = send_report_email(
                        to=email_recipients,
                        location_name=loc_name,
                        year_month=year_month,
                        file_path=output_path,
                        file_name=filename,
                    )
                    if email_result.get("status") == "sent":
                        email_status = "sent"
                        logger.info("%s: email sent to %s", loc_name, email_recipients)

                await supabase.update("batch_run_items", f"id=eq.{item['id']}", {
                    "status": email_status,
                    "row_count": len(processed),
                    "revenue": round(revenue, 2),
                    "file_name": filename,
                    "file_path": output_path,
                    "email_sent_at": datetime.now(timezone.utc).isoformat() if email_status == "sent" else None,
                })
                completed += 1
                logger.info("%s: %d rows, revenue=%.2f", loc_name, len(processed), revenue)

            except Exception as e:
                await supabase.update("batch_run_items", f"id=eq.{item['id']}", {
                    "status": "failed", "error_message": str(e),
                })
                failed += 1
                logger.exception("%s: batch run failed", loc_name)

            # Update progress
            await supabase.update("batch_runs", f"id=eq.{batch_id}", {
                "completed_locations": completed,
                "failed_locations": failed,
            })

        # Done
        await supabase.update("batch_runs", f"id=eq.{batch_id}", {
            "status": "completed",
            "completed_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info("Batch done: %d completed, %d failed", completed, failed)

    except Exception:
        await supabase.update("batch_runs", f"id=eq.{batch_id}", {
            "status": "failed",
        })
        logger.exception("batch run fatal error")


# ─── Location Config ───

@router.get("/locations")
async def list_report_locations():
    """Get locations with report config."""
    items = await supabase.select(
        "locations",
        "select=id,name,station_code,is_report_enabled,email_recipients,electricity_cost,internet_cost,etax,transaction_fee_rate,location_share_rate,share_basis&order=name.asc"
    )
    return {"items": items, "total": len(items)}


class LocationConfigUpdate(BaseModel):
    is_report_enabled: bool | None = None
    email_recipients: list[str] | None = None
    electricity_cost: float | None = None
    internet_cost: float | None = None
    etax: float | None = None
    transaction_fee_rate: float | None = None
    location_share_rate: float | None = None
    share_basis: str | None = None


@router.put("/locations/{location_id}")
async def update_location_config(location_id: str, payload: LocationConfigUpdate):
    data = payload.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(400, "Nothing to update")
    row = await supabase.update("locations", f"id=eq.{location_id}", data)
    return row
