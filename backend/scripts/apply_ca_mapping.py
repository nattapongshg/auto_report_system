"""Seed `locations.ca` from the master SHARGE LOCATIONS xlsx.

The spreadsheet has a 'Locations' sheet with columns:
  Locations (= name), การไฟฟ้า (= 'MEA' | 'PEA'), CA, เลข Meter, ...

We match by `Locations` -> `locations.name` and upsert the CA. Provider is
also stored into `locations.electricity_provider` so the bill-upload flow can
prefer one source when the same CA shows up in both lists (rare but possible).

Usage:
    PYTHONPATH=. ./venv/Scripts/python.exe scripts/apply_ca_mapping.py \
        "c:/auto_report_system/SHARGE LOCATIONS - Copy.xlsx"
"""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

import asyncpg
import openpyxl

from app.config import settings


DEFAULT_XLSX = Path("c:/auto_report_system/SHARGE LOCATIONS - Copy.xlsx")


def normalize_ca(v: object) -> str | None:
    """CA can arrive as int (16880233), unpadded str, or padded str
    ('000017187221'). Canonical form = numeric with no leading zeros — that's
    how SHARGE LOCATIONS stores it and how the PEA/MEA parsers emit it."""
    if v is None or v == "":
        return None
    s = str(v).strip()
    if not s:
        return None
    if s.isdigit():
        s = s.lstrip("0") or "0"
    return s


def parse_xlsx(path: Path) -> list[tuple[str, str, str]]:
    """Returns list of (location_name, provider, ca)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Locations"]
    rows: list[tuple[str, str, str]] = []

    header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(header)}
    name_i = col["Locations"]
    prov_i = col["การไฟฟ้า"]
    ca_i = col["CA"]

    for r in ws.iter_rows(min_row=2, values_only=True):
        name = (r[name_i] or "").strip() if isinstance(r[name_i], str) else r[name_i]
        if not name:
            continue
        prov = (r[prov_i] or "").strip().lower() if r[prov_i] else ""
        ca = normalize_ca(r[ca_i])
        if prov not in ("mea", "pea") or not ca:
            continue
        rows.append((str(name).strip(), prov, ca))
    return rows


async def main() -> None:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    print(f"reading {path}")
    rows = parse_xlsx(path)
    print(f"parsed {len(rows)} CA rows")

    conn = await asyncpg.connect(settings.database_url)
    try:
        # Make sure the column exists (idempotent — no-op if already there).
        await conn.execute("ALTER TABLE locations ADD COLUMN IF NOT EXISTS electricity_provider TEXT")

        updated = 0
        unmatched: list[tuple[str, str, str]] = []
        for name, prov, ca in rows:
            r = await conn.execute(
                """
                UPDATE locations
                   SET ca = $2,
                       electricity_provider = $3
                 WHERE name = $1
                """,
                name, ca, prov,
            )
            if r.endswith(" 1"):
                updated += 1
            else:
                unmatched.append((name, prov, ca))

        print(f"updated: {updated}")
        if unmatched:
            print(f"unmatched ({len(unmatched)}):")
            for name, prov, ca in unmatched:
                print(f"  - [{prov}] {ca} -> {name}")
    finally:
        await conn.close()


if __name__ == "__main__":
    asyncio.run(main())
