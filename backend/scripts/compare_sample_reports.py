"""Build a side-by-side comparison report between the manual-team reference
Excels in sample_report/ and the Q1145 payment-centric output from DB.

Output: `c:/auto_report_system/sample_report_comparison.xlsx`

Columns per location:
  File | Location | Target Revenue | Q1145 Revenue | Diff | Diff%
  Target Rows | Q1145 Invoice Rows | Q1145 Payment Rows
  Q1145 kWh | Q1145 eTax count | Org Rows (fleet) | Priv Rows
  Match (✅/⚠️/❌) | Notes

Run:
    PYTHONPATH=. ./venv/Scripts/python.exe scripts/compare_sample_reports.py
"""
from __future__ import annotations

import asyncio
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.db.payment_rows import load_payment_rows
from app.engine.payment_calc import aggregate_by_location


SAMPLE_DIR = Path("c:/auto_report_system/sample_report")
OUTPUT_XLSX = Path("c:/auto_report_system/sample_report_comparison.xlsx")
SNAPSHOT_ID = "e112307e-7d73-472e-92fe-d3014f95d57e"

# Manual map because file names are abbreviated / different from DB names.
FILE_TO_LOCATION = {
    "BYD CG  Praditmanutham":       "BYD CG Group Praditmanutham",
    "BYD HC BKK Theparak KM.20":    "BYD HC Group Theparak Km.20",
    "BYD Kwang Thai Ekamai-Ramindra": "BYD Kwang Thai Ekamai-Ramindra",
    "BYD Metromobile Rama 3":       "BYD Metromobile Rama 3",
    "BYD SUSCO Beyond BKK Rama 9":  "BYD SUSCO Beyond BKK Rama 9",
    "BYD Harmony Kanlapaphruek":    "BYD Harmony Kanlapaphruek",
    "Denza Susco Sathupradit":      "Denza Susco Sathupradit",
    "BYD Auto Nakhon Si Thammarat": "BYD BD Auto Group Nakhon Si Thammarat",
    "BYD EV-D Sisaket":             "BYD EV-D Sisaket",
    "BYD Phetchaburi EV Phetchaburi": "BYD Phetchaburi EV",
    "BYD SUSCO Beyond AngThong":    "BYD SUSCO Beyond Ang Thong",
    "BYD Yonpiboon Khonkaen":       "BYD Yonpiboon Khonkaen",
    "BYD Arena Motor":              "BYD Arena Motor (Nongbualamphu)",
}


def read_target(path: Path) -> dict:
    """Extract revenue from reference Excel. Uses summary section if present,
    falls back to sum of Payment Amount column."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for sn in wb.sheetnames:
        if "03" in sn and "2026" in sn:
            ws = wb[sn]
            break
    if not ws:
        return {"revenue": 0.0, "rows": 0, "kind": "not_found"}

    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    # Try summary section first
    for r in range(2, ws.max_row + 1):
        for c in range(1, min(ws.max_column + 1, 15)):
            if ws.cell(row=r, column=c).value == "Revenue":
                for dc in (1, 2, 3):
                    rv = ws.cell(row=r, column=c + dc).value
                    if isinstance(rv, (int, float)):
                        # data rows are rows before this with an UUID-ish col 1
                        n = sum(
                            1 for rr in range(2, r)
                            if ws.cell(row=rr, column=1).value
                            and isinstance(ws.cell(row=rr, column=1).value, str)
                            and len(str(ws.cell(row=rr, column=1).value)) >= 30
                        )
                        return {"revenue": float(rv), "rows": n, "kind": "summary"}

    # Fallback: sum Payment Amount
    pay_col = header.get("Payment Amount", 10)
    total = 0.0
    n = 0
    for r in range(2, ws.max_row + 1):
        inv = ws.cell(row=r, column=1).value
        if not inv or not isinstance(inv, str) or len(inv) < 30:
            continue
        pay = ws.cell(row=r, column=pay_col).value
        if isinstance(pay, (int, float)):
            total += pay
            n += 1
    return {"revenue": total, "rows": n, "kind": "sum_payment"}


async def build_row(fp: Path, loc_name: str, agg: dict) -> dict:
    target = read_target(fp)
    bucket = agg.get(loc_name, {})
    rev_q = float(bucket.get("revenue", 0))
    rev_t = target["revenue"]
    diff = rev_q - rev_t
    diff_pct = (diff / rev_t * 100) if rev_t > 0 else None

    # Load DB payment rows for refund / privilege counts
    rows, cols = await load_payment_rows(SNAPSHOT_ID, location_name=loc_name)
    idx = {n: i for i, n in enumerate(cols)}
    full_refund = sum(1 for r in rows if r[idx["refund_type"]] == "full")
    partial_refund = sum(1 for r in rows if r[idx["refund_type"]] == "partial")
    org_rows = sum(1 for r in rows if r[idx["invoice_status"]] == "billed_to_organization")
    priv_rows = sum(1 for r in rows if r[idx["privilege_program_name"]])

    if abs(diff) < 0.5:
        match = "✅ match"
    elif abs(diff) < 100:
        match = "⚠️ small"
    else:
        match = "❌ miss"

    return {
        "file": fp.name,
        "location": loc_name,
        "target_rev": rev_t,
        "q1145_rev": rev_q,
        "diff": diff,
        "diff_pct": diff_pct,
        "target_rows": target["rows"],
        "q1145_invoice_rows": bucket.get("invoice_rows", 0),
        "q1145_payment_rows": bucket.get("payment_rows", 0),
        "q1145_kwh": bucket.get("kwh", 0),
        "q1145_etax": bucket.get("etax_count", 0),
        "full_refund": full_refund,
        "partial_refund": partial_refund,
        "org_rows": org_rows,
        "priv_rows": priv_rows,
        "target_kind": target["kind"],
        "match": match,
    }


def write_report(rows: list[dict], out_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    headers = [
        "File", "Location (DB)", "Match",
        "Target Revenue", "Q1145 Revenue", "Diff", "Diff %",
        "Target Rows", "Q1145 Inv Rows", "Q1145 Pay Rows",
        "Q1145 kWh", "Q1145 eTax",
        "Full Refund", "Partial Refund", "Org Rows", "Priv Rows",
        "Target Format",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1a1a2e")
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, 2):
        vals = [
            r["file"], r["location"], r["match"],
            r["target_rev"], r["q1145_rev"], r["diff"],
            f"{r['diff_pct']:+.3f}%" if r["diff_pct"] is not None else "—",
            r["target_rows"], r["q1145_invoice_rows"], r["q1145_payment_rows"],
            round(r["q1145_kwh"], 2), r["q1145_etax"],
            r["full_refund"], r["partial_refund"], r["org_rows"], r["priv_rows"],
            r["target_kind"],
        ]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=j, value=v)
            if j in (4, 5, 6, 11):
                cell.number_format = "#,##0.00"
            # Colour-code match column
            if j == 3:
                if "✅" in str(v):
                    cell.fill = PatternFill("solid", fgColor="d4edda")
                elif "⚠️" in str(v):
                    cell.fill = PatternFill("solid", fgColor="fff3cd")
                elif "❌" in str(v):
                    cell.fill = PatternFill("solid", fgColor="f8d7da")
            # Colour-code diff
            if j == 6 and isinstance(v, (int, float)):
                if abs(v) < 0.5:
                    cell.fill = PatternFill("solid", fgColor="d4edda")
                elif abs(v) < 100:
                    cell.fill = PatternFill("solid", fgColor="fff3cd")
                else:
                    cell.fill = PatternFill("solid", fgColor="f8d7da")

    # Totals row
    n = len(rows) + 1
    tot_row = n + 1
    ws.cell(row=tot_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=tot_row, column=4, value=sum(r["target_rev"] for r in rows)).number_format = "#,##0.00"
    ws.cell(row=tot_row, column=5, value=sum(r["q1145_rev"] for r in rows)).number_format = "#,##0.00"
    ws.cell(row=tot_row, column=6, value=sum(r["diff"] for r in rows)).number_format = "#,##0.00"
    ws.cell(row=tot_row, column=8, value=sum(r["target_rows"] for r in rows))
    ws.cell(row=tot_row, column=9, value=sum(r["q1145_invoice_rows"] for r in rows))
    ws.cell(row=tot_row, column=10, value=sum(r["q1145_payment_rows"] for r in rows))
    ws.cell(row=tot_row, column=11, value=round(sum(r["q1145_kwh"] for r in rows), 2))
    ws.cell(row=tot_row, column=12, value=sum(r["q1145_etax"] for r in rows))
    ws.cell(row=tot_row, column=13, value=sum(r["full_refund"] for r in rows))
    ws.cell(row=tot_row, column=14, value=sum(r["partial_refund"] for r in rows))
    ws.cell(row=tot_row, column=15, value=sum(r["org_rows"] for r in rows))
    ws.cell(row=tot_row, column=16, value=sum(r["priv_rows"] for r in rows))
    for j in range(1, 18):
        ws.cell(row=tot_row, column=j).font = Font(bold=True)
        ws.cell(row=tot_row, column=j).fill = PatternFill("solid", fgColor="f0f0f0")

    # Summary stats section
    s_row = tot_row + 3
    ws.cell(row=s_row, column=1, value="Summary").font = Font(bold=True, size=14)
    matches = sum(1 for r in rows if "✅" in r["match"])
    small = sum(1 for r in rows if "⚠️" in r["match"])
    miss = sum(1 for r in rows if "❌" in r["match"])
    ws.cell(row=s_row + 1, column=1, value=f"✅ Perfect match (diff < 0.5): {matches}/{len(rows)}")
    ws.cell(row=s_row + 2, column=1, value=f"⚠️ Small diff (< 100): {small}/{len(rows)}")
    ws.cell(row=s_row + 3, column=1, value=f"❌ Big miss (>= 100): {miss}/{len(rows)}")

    # Auto-size columns
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, tot_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 2

    wb.save(out_path)


async def main():
    print("Aggregating Q1145 across all locations...")
    agg = await aggregate_by_location(SNAPSHOT_ID)
    print(f"  → {len(agg)} locations in snapshot\n")

    rows = []
    for fp in sorted(SAMPLE_DIR.iterdir()):
        if not fp.name.endswith(".xlsx") or fp.name.startswith("~"):
            continue
        loc = None
        for frag, name in FILE_TO_LOCATION.items():
            if frag in fp.name:
                loc = name
                break
        if not loc:
            print(f"  ??? {fp.name} — no mapping")
            continue
        row = await build_row(fp, loc, agg)
        rows.append(row)
        print(f"  {row['match']} {fp.name[:55]:55} diff={row['diff']:+,.2f}")

    write_report(rows, OUTPUT_XLSX)
    print(f"\nReport saved: {OUTPUT_XLSX}")


if __name__ == "__main__":
    asyncio.run(main())
